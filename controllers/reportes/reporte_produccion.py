from collections import defaultdict
from datetime import date, datetime
import os
from typing import Dict, Any, List, Tuple

from flask import session, current_app

from models.db import get_connection
from utils.financiamiento_grupal_reportes import enrich_rows_with_fg_metadata
from utils.rbac import Roles


MONTH_NAMES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Setiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}


def get_reporte_produccion_filters() -> Dict[str, Any]:
    from controllers.compania import get_aseguradoras
    from controllers.ramos import get_ramos
    from controllers.subagente import get_subagentes_abreviaciones
    from controllers.ejecutivos import get_ejecutivos
    from controllers.maestros.usuarios import get_usuarios

    return {
        "companias": get_aseguradoras() or [],
        "ramos": get_ramos() or [],
        "subagentes": get_subagentes_abreviaciones() or [],
        "ejecutivos": get_ejecutivos() or [],
        "usuarios": get_usuarios() or [],
    }


def _build_filters(filters: Dict[str, Any]) -> Tuple[str, List[Any]]:
    sql_filters = []
    params: List[Any] = []

    sql_filters.append(
        "COALESCE(NULLIF(TRIM(REPLACE(CONVERT(p.activo USING latin1), _latin1 0xA0, ' ')), ''), '0') = '1'"
    )
    sql_filters.append(
        "COALESCE(NULLIF(TRIM(REPLACE(CONVERT(p.anulado USING latin1), _latin1 0xA0, ' ')), ''), '0') = '0'"
    )
    sql_filters.append("COALESCE(p.prima_anulada, 0) = 0")
    sql_filters.append(
        "("
        "p.estado IS NULL "
        "OR TRIM(CONVERT(p.estado USING utf8mb4)) = '' "
        "OR UPPER(TRIM(CONVERT(p.estado USING utf8mb4))) NOT LIKE 'ANULAD%'"
        ")"
    )

    vig_desde = filters.get("vig_desde")
    vig_hasta = filters.get("vig_hasta")

    if vig_desde:
        sql_filters.append("p.vig_desde >= %s")
        params.append(vig_desde)

    if vig_hasta:
        sql_filters.append("p.vig_desde <= %s")
        params.append(vig_hasta)

    if filters.get("cia"):
        sql_filters.append("p.cia = %s")
        params.append(filters["cia"])

    if filters.get("ramo"):
        sql_filters.append("p.ramo = %s")
        params.append(filters["ramo"])

    if filters.get("sub_agente"):
        sql_filters.append("p.sub_agente = %s")
        params.append(filters["sub_agente"])

    if filters.get("ejecutivo"):
        sql_filters.append("p.ejecutivo = %s")
        params.append(filters["ejecutivo"])

    if filters.get("moneda"):
        mon = (filters.get("moneda") or "").strip().upper()
        if mon in {"S/", "S/.", "SOLES", "PEN"}:
            sql_filters.append(
                "(UPPER(TRIM(p.moneda)) LIKE 'S/%' OR UPPER(TRIM(p.moneda)) IN ('SOLES','PEN'))"
            )
        elif mon in {"US$", "USD", "$", "DOLARES", "DÓLARES", "DOLAR"}:
            sql_filters.append(
                "(UPPER(TRIM(p.moneda)) IN ('US$','USD','$','DOLARES','DÓLARES','DOLAR'))"
            )
        else:
            sql_filters.append("p.moneda = %s")
            params.append(filters["moneda"])

    # Filtro por rol: sub agente solo ve sus pólizas
    role = session.get("role_name")
    user = session.get("user")

    if role == Roles.SUB_AGENTE and user:
        sql_filters.append(
            "("
            "p.sub_agente = %s "
            "OR p.sub_agente = (SELECT COALESCE(NULLIF(TRIM(nombre), ''), username) FROM usuarios WHERE username = %s LIMIT 1) "
            "OR p.usuario_registro = %s "
            "OR p.usuario_registro = (SELECT COALESCE(NULLIF(TRIM(nombre), ''), username) FROM usuarios WHERE username = %s LIMIT 1)"
            ")"
        )
        params.extend([user, user, user, user])

    where_clause = ""
    if sql_filters:
        where_clause = " WHERE " + " AND ".join(sql_filters)

    return where_clause, params


def get_reporte_produccion_rows(filters: Dict[str, Any], limit: int = 1000) -> List[Dict[str, Any]]:
    conn = get_connection()
    try:
        cursor = conn.cursor(dictionary=True)

        base_sql = """
            SELECT
                p.idPoliza AS idPoliza,
                COALESCE(CAST(AES_DECRYPT(FROM_BASE64(c.numero_documento), @SIS_KEY) AS CHAR), c.numero_documento) AS ruc,
                COALESCE(CAST(AES_DECRYPT(FROM_BASE64(c.razon_social), @SIS_KEY) AS CHAR), c.razon_social) AS contratante,
                COALESCE(CAST(AES_DECRYPT(FROM_BASE64(c.direccion), @SIS_KEY) AS CHAR), c.direccion) AS direccion_contratante,
                COALESCE(CAST(AES_DECRYPT(FROM_BASE64(p.asegurado), @SIS_KEY) AS CHAR), p.asegurado) AS asegurado,
                p.cia,
                p.ramo AS ram,
                p.ramos_producto AS prod,
                COALESCE(CAST(AES_DECRYPT(FROM_BASE64(p.poliza), @SIS_KEY) AS CHAR), p.poliza) AS poliza,
                p.tipo_doc AS td,
                COALESCE(CAST(AES_DECRYPT(FROM_BASE64(p.recibo), @SIS_KEY) AS CHAR), p.recibo) AS aviso_cob,
                '' AS estado_comision,
                p.vig_desde AS ini_vig,
                p.vig_hasta AS fin_vig,
                p.moneda AS mon,
                COALESCE(p.prima_neta, 0) AS prim_neta,
                COALESCE(p.prima_comercial_igv) AS prim_total,
                p.porc_compania AS porc_cia,
                p.imp_compania AS comision_cia,
                p.sub_agente AS sagt,
                p.porc_subagente AS porc_sagt,
                p.imp_subagente AS comision_sagt,
                NULL AS fpago_sagt,
                NULL AS comprobante_sagt,
                p.motivo,
                c.departamento AS ciudad,
                NULL AS factura_comision,
                p.ejecutivo,
                p.asegurada AS breve_descripcion,
                p.usuario_registro AS usuario,
                DATE(p.creado_en) AS f_reg
            FROM polizas p
            INNER JOIN clientes c ON c.idCliente = p.cliente_id
        """

        where_clause, params = _build_filters(filters)
        
        order_clause = " ORDER BY p.fecha_emision DESC, p.creado_en DESC"
        if limit:
            order_clause += f" LIMIT {limit}"

        sql = base_sql + where_clause + order_clause

        cursor.execute(sql, tuple(params))
        rows = cursor.fetchall() or []
        enrich_rows_with_fg_metadata(rows, poliza_id_keys=("idPoliza", "poliza_id"))
        
        # Post-process to format dates as strings
        for row in rows:
            if row.get('ini_vig'):
                row['ini_vig'] = str(row['ini_vig'])
            if row.get('fin_vig'):
                row['fin_vig'] = str(row['fin_vig'])
            if row.get('f_reg'):
                row['f_reg'] = str(row['f_reg'])

        return rows
    finally:
        conn.close()


def _get_export_headers() -> List[str]:
    return [
        "RUC",
        "CONTRATANTE",
        "DIRECCION DEL CONTRATANTE",
        "ASEGURADO",
        "CIA",
        "RAM",
        "PROD",
        "POLIZA",
        "TD",
        "AVISO COB",
        "ESTADO COMISION",
        "INI.VIG",
        "FIN.VIG",
        "MON",
        "PRIM.NETA",
        "PRIM.TOTAL",
        "% CIA",
        "COMISION CIA",
        "SAGT",
        "% SAGT",
        "COMISION SAGT",
        "F.PAGO SAGT",
        "COMPROBANTE SAGT",
        "MOTIVO",
        "CIUDAD",
        "FACTURA COMISION",
        "EJECUTIVO",
        "BREVE DESCRIPCION",
        "Usuario",
        "F.REG",
    ]


def _build_export_table_rows(rows: List[Dict[str, Any]]) -> List[List[Any]]:
    table_rows: List[List[Any]] = []
    for r in rows:
        table_rows.append(
            [
                r.get("idPoliza"),
                r.get("ruc") or "",
                r.get("contratante") or "",
                r.get("direccion_contratante") or "",
                r.get("asegurado") or "",
                r.get("cia") or "",
                r.get("ram") or "",
                r.get("prod") or "",
                r.get("poliza") or "",
                r.get("td") or "",
                r.get("aviso_cob") or "",
                r.get("estado_comision") or "",
                r.get("ini_vig"),
                r.get("fin_vig"),
                r.get("mon") or "",
                float(r.get("prim_neta") or 0),
                float(r.get("prim_total") or 0),
                float(r.get("porc_cia") or 0),
                float(r.get("comision_cia") or 0),
                r.get("sagt") or "",
                float(r.get("porc_sagt") or 0),
                float(r.get("comision_sagt") or 0),
                r.get("fpago_sagt"),
                r.get("comprobante_sagt") or "",
                r.get("motivo") or "",
                r.get("ciudad") or "",
                r.get("factura_comision") or "",
                r.get("ejecutivo") or "",
                r.get("breve_descripcion") or "",
                r.get("usuario") or "",
                r.get("f_reg"),
            ]
        )
    return table_rows


def _resolve_exports_dir() -> str:
    try:
        upload_folder = current_app.config.get("UPLOAD_FOLDER")
    except Exception:
        base_dir = os.path.dirname(os.path.dirname(__file__))
        upload_folder = os.path.join(base_dir, "..", "uploads")

    if not upload_folder:
        base_dir = os.path.dirname(os.path.dirname(__file__))
        upload_folder = os.path.join(base_dir, "..", "uploads")

    exports_dir = os.path.join(upload_folder, "exports")
    os.makedirs(exports_dir, exist_ok=True)
    return exports_dir


def _safe_float(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _calc_productivity(prima_total: float, total_comision: float) -> float:
    if not prima_total:
        return 0.0
    return (total_comision / prima_total) * 100.0


def _parse_report_date(value: Any) -> date | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def _build_dashboard_summary(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    daily = defaultdict(
        lambda: {
            "cantidad": 0,
            "prima_neta": 0.0,
            "prima_total": 0.0,
            "comision_cia": 0.0,
            "comision_sagt": 0.0,
        }
    )
    monthly = defaultdict(
        lambda: {
            "cantidad": 0,
            "prima_neta": 0.0,
            "prima_total": 0.0,
            "comision_cia": 0.0,
            "comision_sagt": 0.0,
        }
    )
    yearly = defaultdict(
        lambda: {
            "cantidad": 0,
            "prima_neta": 0.0,
            "prima_total": 0.0,
            "comision_cia": 0.0,
            "comision_sagt": 0.0,
            "months": defaultdict(float),
        }
    )
    ramo_totals = defaultdict(float)
    ramo_records_map = defaultdict(
        lambda: {
            "cantidad": 0,
            "prima_neta": 0.0,
            "prima_total": 0.0,
            "comision_cia": 0.0,
            "comision_sagt": 0.0,
        }
    )
    contratante_totals = defaultdict(
        lambda: {
            "cantidad": 0,
            "prima_neta": 0.0,
            "prima_total": 0.0,
            "comision_cia": 0.0,
            "comision_sagt": 0.0,
            "ramos": set(),
        }
    )

    total_cantidad = 0
    total_prima_neta = 0.0
    total_prima_total = 0.0
    total_comision_cia = 0.0
    total_comision_sagt = 0.0

    for row in rows:
        row_date = _parse_report_date(row.get("ini_vig") or row.get("f_reg"))
        if not row_date:
            continue

        year = row_date.year
        month = row_date.month
        prima_neta = _safe_float(row.get("prim_neta"))
        prima_total = _safe_float(row.get("prim_total"))
        comision_cia = _safe_float(row.get("comision_cia"))
        comision_sagt = _safe_float(row.get("comision_sagt"))
        total_comision = comision_cia + comision_sagt
        contratante = (row.get("contratante") or "SIN CONTRATANTE").strip() or "SIN CONTRATANTE"

        daily[row_date]["cantidad"] += 1
        daily[row_date]["prima_neta"] += prima_neta
        daily[row_date]["prima_total"] += prima_total
        daily[row_date]["comision_cia"] += comision_cia
        daily[row_date]["comision_sagt"] += comision_sagt

        monthly_key = (year, month)
        monthly[monthly_key]["cantidad"] += 1
        monthly[monthly_key]["prima_neta"] += prima_neta
        monthly[monthly_key]["prima_total"] += prima_total
        monthly[monthly_key]["comision_cia"] += comision_cia
        monthly[monthly_key]["comision_sagt"] += comision_sagt

        yearly[year]["cantidad"] += 1
        yearly[year]["prima_neta"] += prima_neta
        yearly[year]["prima_total"] += prima_total
        yearly[year]["comision_cia"] += comision_cia
        yearly[year]["comision_sagt"] += comision_sagt
        yearly[year]["months"][month] += prima_total

        ramo = (row.get("ram") or "SIN RAMO").strip() or "SIN RAMO"
        ramo_totals[ramo] += prima_total
        ramo_records_map[ramo]["cantidad"] += 1
        ramo_records_map[ramo]["prima_neta"] += prima_neta
        ramo_records_map[ramo]["prima_total"] += prima_total
        ramo_records_map[ramo]["comision_cia"] += comision_cia
        ramo_records_map[ramo]["comision_sagt"] += comision_sagt

        contratante_totals[contratante]["cantidad"] += 1
        contratante_totals[contratante]["prima_neta"] += prima_neta
        contratante_totals[contratante]["prima_total"] += prima_total
        contratante_totals[contratante]["comision_cia"] += comision_cia
        contratante_totals[contratante]["comision_sagt"] += comision_sagt
        contratante_totals[contratante]["ramos"].add(ramo)

        total_cantidad += 1
        total_prima_neta += prima_neta
        total_prima_total += prima_total
        total_comision_cia += comision_cia
        total_comision_sagt += comision_sagt

    daily_records: List[Dict[str, Any]] = []
    for item_date, values in sorted(daily.items()):
        total_comision = values["comision_cia"] + values["comision_sagt"]
        daily_records.append(
            {
                "fecha": item_date.isoformat(),
                "anio": item_date.year,
                "mes_numero": item_date.month,
                "mes": MONTH_NAMES[item_date.month],
                "dia": item_date.day,
                "cantidad": values["cantidad"],
                "prima_neta": values["prima_neta"],
                "prima_total": values["prima_total"],
                "comision_cia": values["comision_cia"],
                "comision_sagt": values["comision_sagt"],
                "total_comision": total_comision,
                "productividad_pct": _calc_productivity(values["prima_total"], total_comision),
            }
        )

    monthly_records: List[Dict[str, Any]] = []
    for (year, month), values in sorted(monthly.items()):
        total_comision = values["comision_cia"] + values["comision_sagt"]
        monthly_records.append(
            {
                "periodo": f"{year}-{month:02d}",
                "anio": year,
                "mes_numero": month,
                "mes": MONTH_NAMES[month],
                "cantidad": values["cantidad"],
                "prima_neta": values["prima_neta"],
                "prima_total": values["prima_total"],
                "comision_cia": values["comision_cia"],
                "comision_sagt": values["comision_sagt"],
                "total_comision": total_comision,
                "productividad_pct": _calc_productivity(values["prima_total"], total_comision),
            }
        )

    yearly_records: List[Dict[str, Any]] = []
    best_months_by_year: List[Dict[str, Any]] = []
    for year in sorted(yearly.keys()):
        months_map = yearly[year]["months"]
        total_comision = yearly[year]["comision_cia"] + yearly[year]["comision_sagt"]
        best_month_num = None
        best_month_total = 0.0
        if months_map:
            best_month_num, best_month_total = max(months_map.items(), key=lambda item: item[1])

        yearly_records.append(
            {
                "anio": year,
                "cantidad": yearly[year]["cantidad"],
                "prima_neta": yearly[year]["prima_neta"],
                "prima_total": yearly[year]["prima_total"],
                "comision_cia": yearly[year]["comision_cia"],
                "comision_sagt": yearly[year]["comision_sagt"],
                "total_comision": total_comision,
                "productividad_pct": _calc_productivity(yearly[year]["prima_total"], total_comision),
                "meses": [months_map.get(month, 0.0) for month in range(1, 13)],
            }
        )

        best_months_by_year.append(
            {
                "anio": year,
                "mes": MONTH_NAMES.get(best_month_num, ""),
                "prima_total": best_month_total,
            }
        )

    best_month = None
    if monthly_records:
        best_month = max(monthly_records, key=lambda item: (item["prima_total"], item["cantidad"]))

    top_ramos = [
        {"ramo": ramo, "prima_total": total}
        for ramo, total in sorted(ramo_totals.items(), key=lambda item: item[1], reverse=True)[:10]
    ]

    ramo_records: List[Dict[str, Any]] = []
    for ramo, values in sorted(ramo_records_map.items(), key=lambda item: item[1]["prima_total"], reverse=True):
        total_comision = values["comision_cia"] + values["comision_sagt"]
        ramo_records.append(
            {
                "ramo": ramo,
                "cantidad": values["cantidad"],
                "prima_neta": values["prima_neta"],
                "prima_total": values["prima_total"],
                "comision_cia": values["comision_cia"],
                "comision_sagt": values["comision_sagt"],
                "total_comision": total_comision,
                "productividad_pct": _calc_productivity(values["prima_total"], total_comision),
            }
        )

    contratante_records: List[Dict[str, Any]] = []
    for contratante, values in sorted(contratante_totals.items(), key=lambda item: item[1]["prima_total"], reverse=True):
        total_comision = values["comision_cia"] + values["comision_sagt"]
        ramos = sorted(values["ramos"])
        contratante_records.append(
            {
                "contratante": contratante,
                "cantidad": values["cantidad"],
                "cantidad_ramos": len(ramos),
                "ramos": " | ".join(ramos),
                "prima_neta": values["prima_neta"],
                "prima_total": values["prima_total"],
                "comision_cia": values["comision_cia"],
                "comision_sagt": values["comision_sagt"],
                "total_comision": total_comision,
                "productividad_pct": _calc_productivity(values["prima_total"], total_comision),
            }
        )

    top_contratantes = contratante_records[:15]
    low_contratantes = sorted(contratante_records, key=lambda item: (item["prima_total"], item["cantidad"]))[:15]

    best_day = None
    if daily_records:
        best_day = max(daily_records, key=lambda item: (item["prima_total"], item["cantidad"]))

    total_comision = total_comision_cia + total_comision_sagt

    return {
        "totales": {
            "cantidad": total_cantidad,
            "prima_neta": total_prima_neta,
            "prima_total": total_prima_total,
            "comision_cia": total_comision_cia,
            "comision_sagt": total_comision_sagt,
            "total_comision": total_comision,
            "productividad_pct": _calc_productivity(total_prima_total, total_comision),
        },
        "daily_records": daily_records,
        "monthly_records": monthly_records,
        "yearly_records": yearly_records,
        "ramo_records": ramo_records,
        "contratante_records": contratante_records,
        "best_month": best_month,
        "best_day": best_day,
        "best_months_by_year": best_months_by_year,
        "top_ramos": top_ramos,
        "top_contratantes": top_contratantes,
        "low_contratantes": low_contratantes,
    }


def _build_year_month_pivots(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    years_set = set()
    ramo_map = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    ramo_totals = defaultdict(float)
    contratante_map = defaultdict(
        lambda: {
            "ramos": defaultdict(lambda: defaultdict(lambda: defaultdict(float))),
            "totals": defaultdict(lambda: defaultdict(float)),
            "grand_total": 0.0,
        }
    )

    for row in rows:
        row_date = _parse_report_date(row.get("ini_vig") or row.get("f_reg"))
        if not row_date:
            continue

        year = row_date.year
        month = row_date.month
        years_set.add(year)

        value = _safe_float(row.get("prim_neta"))
        contratante = (row.get("contratante") or "SIN CONTRATANTE").strip() or "SIN CONTRATANTE"
        ramo = (row.get("ram") or "SIN RAMO").strip() or "SIN RAMO"

        ramo_map[ramo][year][month] += value
        ramo_totals[ramo] += value

        contratante_map[contratante]["ramos"][ramo][year][month] += value
        contratante_map[contratante]["totals"][year][month] += value
        contratante_map[contratante]["grand_total"] += value

    years = sorted(years_set)

    ramo_rows: List[Dict[str, Any]] = []
    for ramo, year_map in sorted(ramo_map.items(), key=lambda item: ramo_totals[item[0]], reverse=True):
        ramo_rows.append(
            {
                "ramo": ramo,
                "grand_total": ramo_totals[ramo],
                "years": {
                    year: {month: year_map[year].get(month, 0.0) for month in range(1, 13)}
                    for year in years
                },
            }
        )

    contratante_rows: List[Dict[str, Any]] = []
    for contratante, data in sorted(contratante_map.items(), key=lambda item: item[1]["grand_total"], reverse=True):
        ramo_items = []
        for ramo, year_map in sorted(
            data["ramos"].items(),
            key=lambda item: sum(item[1][year].get(month, 0.0) for year in item[1] for month in item[1][year]),
            reverse=True,
        ):
            ramo_items.append(
                {
                    "ramo": ramo,
                    "years": {
                        year: {month: year_map[year].get(month, 0.0) for month in range(1, 13)}
                        for year in years
                    },
                    "grand_total": sum(year_map[year].get(month, 0.0) for year in year_map for month in year_map[year]),
                }
            )

        contratante_rows.append(
            {
                "contratante": contratante,
                "grand_total": data["grand_total"],
                "years": {
                    year: {month: data["totals"][year].get(month, 0.0) for month in range(1, 13)}
                    for year in years
                },
                "ramos": ramo_items,
            }
        )

    return {
        "years": years,
        "ramo_rows": ramo_rows,
        "contratante_rows": contratante_rows,
    }


def export_reporte_produccion(filters: Dict[str, Any]) -> Tuple[str, str]:
    # Para exportación, traemos todo (o un límite alto)
    rows = get_reporte_produccion_rows(filters, limit=None)
    headers = _get_export_headers()
    table_rows = _build_export_table_rows(rows)
    exports_dir = _resolve_exports_dir()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"reporte_produccion_{ts}.xlsx"
    filepath = os.path.join(exports_dir, filename)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Produccion"

    title = "REPORTE DE PRODUCCIÓN"
    vig_desde = filters.get("vig_desde") or ""
    vig_hasta = filters.get("vig_hasta") or ""
    if vig_desde and vig_hasta:
        title = f"{title} — {vig_desde} a {vig_hasta}"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F59A3")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    header_fill = PatternFill("solid", fgColor="399AD6")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 22

    money_cols = {15, 16, 18, 21}
    percent_cols = {17, 20}

    for row_idx, row in enumerate(table_rows, start=3):
        fill = PatternFill("solid", fgColor="F4F8FF") if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, value in enumerate(row, start=1):
            if col_idx == 1:
                continue
            excel_col_idx = col_idx - 1
            cell = ws.cell(row=row_idx, column=excel_col_idx, value=value)
            cell.border = border
            cell.fill = fill
            cell.font = Font(size=9)
            if excel_col_idx in money_cols:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif excel_col_idx in percent_cols:
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if excel_col_idx == 8:
                fg_meta = rows[row_idx - 3] if row_idx - 3 < len(rows) else {}
                if fg_meta.get("es_financiamiento_grupal"):
                    cell.font = Font(size=9, bold=True, color="7A3DB8")

    total_row = len(table_rows) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=9)
    total_cols = [15, 16, 18, 21]
    for col in total_cols:
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True, size=9)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if table_rows:
            col_letter = get_column_letter(col)
            cell.value = f"=SUBTOTAL(109,{col_letter}3:{col_letter}{total_row-1})"
        else:
            cell.value = 0

    for c in [15, 16, 18, 21]:
        ws.cell(row=total_row, column=c).number_format = '#,##0.00'
        ws.cell(row=total_row, column=c).alignment = Alignment(horizontal="right", vertical="center")

    col_widths = [
        12, 28, 28, 28, 16, 18, 18, 16, 6, 16,
        14, 12, 12, 10, 12, 12, 8, 12, 16, 8,
        12, 12, 18, 18, 14, 16, 16, 18, 16, 12
    ]
    for i, w in enumerate(col_widths[:len(headers)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(filepath)

    return filepath, filename


def export_reporte_produccion_pro(filters: Dict[str, Any]) -> Tuple[str, str]:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    rows = get_reporte_produccion_rows(filters, limit=None)
    headers = _get_export_headers()
    table_rows = _build_export_table_rows(rows)
    dashboard = _build_dashboard_summary(rows)
    pivots = _build_year_month_pivots(rows)
    exports_dir = _resolve_exports_dir()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"reporte_produccion_pro_{ts}.xlsx"
    filepath = os.path.join(exports_dir, filename)

    wb = Workbook()
    ws_dashboard = wb.active
    ws_dashboard.title = "Dashboard"
    ws_daily = wb.create_sheet("Productividad Diaria")
    ws_monthly = wb.create_sheet("Resumen Mensual")
    ws_yearly = wb.create_sheet("Resumen Anual")
    ws_ramo_matrix = wb.create_sheet("Ramos Anio-Mes")
    ws_contratante_matrix = wb.create_sheet("Contratantes Anio-Mes")
    ws_ramo = wb.create_sheet("Ramos")
    ws_contratante = wb.create_sheet("Contratantes")
    ws_productivity = wb.create_sheet("Productividad Anual")
    ws_detail = wb.create_sheet("Detalle")

    primary_fill = PatternFill("solid", fgColor="1F4E78")
    secondary_fill = PatternFill("solid", fgColor="D9EAF7")
    accent_fill = PatternFill("solid", fgColor="E2F0D9")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_format = '#,##0.00'

    def style_table_header(sheet, row_num: int, max_col: int) -> None:
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row_num, column=col)
            cell.fill = primary_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    def add_excel_table(sheet, start_row: int, end_row: int, end_col: int, table_name: str) -> None:
        if end_row <= start_row:
            return
        ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    month_short_names = {
        1: "Ene",
        2: "Feb",
        3: "Mar",
        4: "Abr",
        5: "May",
        6: "Jun",
        7: "Jul",
        8: "Ago",
        9: "Set",
        10: "Oct",
        11: "Nov",
        12: "Dic",
    }

    def style_matrix_cell(cell, is_number: bool = False, fill=None, bold: bool = False) -> None:
        cell.border = border
        if fill is not None:
            cell.fill = fill
        if bold:
            cell.font = bold_font
        if is_number:
            if cell.value in (None, ""):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.number_format = money_format
                cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def write_year_month_matrix_header(sheet, fixed_headers: List[str], years: List[int]) -> int:
        year_list = years or [datetime.now().year]
        for idx, header in enumerate(fixed_headers, start=1):
            sheet.merge_cells(start_row=1, start_column=idx, end_row=2, end_column=idx)
            cell = sheet.cell(row=1, column=idx, value=header)
            cell.fill = primary_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            sheet.cell(row=2, column=idx).fill = primary_fill
            sheet.cell(row=2, column=idx).border = border

        col = len(fixed_headers) + 1
        for year in year_list:
            start_col = col
            end_col = col + 12
            sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            year_cell = sheet.cell(row=1, column=start_col, value=str(year))
            year_cell.fill = primary_fill
            year_cell.font = white_font
            year_cell.alignment = Alignment(horizontal="center", vertical="center")
            year_cell.border = border
            for month in range(1, 13):
                month_cell = sheet.cell(row=2, column=col, value=month_short_names[month])
                month_cell.fill = primary_fill
                month_cell.font = white_font
                month_cell.alignment = Alignment(horizontal="center", vertical="center")
                month_cell.border = border
                col += 1
            total_cell = sheet.cell(row=2, column=col, value=f"Total {year}")
            total_cell.fill = primary_fill
            total_cell.font = white_font
            total_cell.alignment = Alignment(horizontal="center", vertical="center")
            total_cell.border = border
            col += 1
        return col - 1

    def write_year_values(sheet, row_num: int, start_col: int, years: List[int], year_values: Dict[int, Dict[int, float]], fill=None, bold: bool = False) -> None:
        col = start_col
        for year in years:
            total_year = 0.0
            month_map = year_values.get(year, {})
            for month in range(1, 13):
                value = month_map.get(month, 0.0)
                total_year += value
                cell = sheet.cell(row=row_num, column=col, value=value if value else "")
                style_matrix_cell(cell, is_number=True, fill=fill, bold=bold)
                col += 1
            total_cell = sheet.cell(row=row_num, column=col, value=total_year if total_year else "")
            style_matrix_cell(total_cell, is_number=True, fill=fill, bold=bold)
            col += 1

    def render_ramo_matrix(sheet) -> None:
        years = pivots["years"] or [datetime.now().year]
        last_col = write_year_month_matrix_header(sheet, ["RAMO"], years)
        row_num = 3
        for item in pivots["ramo_rows"]:
            label_cell = sheet.cell(row=row_num, column=1, value=item["ramo"])
            style_matrix_cell(label_cell)
            write_year_values(sheet, row_num, 2, years, item["years"])
            row_num += 1

        total_map = {
            year: {
                month: sum(item["years"].get(year, {}).get(month, 0.0) for item in pivots["ramo_rows"])
                for month in range(1, 13)
            }
            for year in years
        }
        total_label = sheet.cell(row=row_num, column=1, value="Total general")
        style_matrix_cell(total_label, fill=secondary_fill, bold=True)
        write_year_values(sheet, row_num, 2, years, total_map, fill=secondary_fill, bold=True)
        sheet.freeze_panes = "B3"
        sheet.auto_filter.ref = f"A2:{get_column_letter(last_col)}{row_num}"

    def render_contratante_matrix(sheet) -> None:
        years = pivots["years"] or [datetime.now().year]
        last_col = write_year_month_matrix_header(sheet, ["CONTRATANTE", "RAMO"], years)
        row_num = 3
        for item in pivots["contratante_rows"]:
            first_row = True
            for ramo_item in item["ramos"]:
                contratante_value = item["contratante"] if first_row else ""
                contratante_cell = sheet.cell(row=row_num, column=1, value=contratante_value)
                style_matrix_cell(contratante_cell)
                ramo_cell = sheet.cell(row=row_num, column=2, value=ramo_item["ramo"])
                style_matrix_cell(ramo_cell)
                write_year_values(sheet, row_num, 3, years, ramo_item["years"])
                row_num += 1
                first_row = False

            total_cell = sheet.cell(row=row_num, column=1, value=f"Total {item['contratante']}")
            style_matrix_cell(total_cell, fill=secondary_fill, bold=True)
            total_ramo_cell = sheet.cell(row=row_num, column=2, value="")
            style_matrix_cell(total_ramo_cell, fill=secondary_fill, bold=True)
            write_year_values(sheet, row_num, 3, years, item["years"], fill=secondary_fill, bold=True)
            row_num += 1

        sheet.freeze_panes = "C3"
        sheet.auto_filter.ref = f"A2:{get_column_letter(last_col)}{row_num - 1}"

    title = "REPORTE DE PRODUCCION PRO"
    vig_desde = filters.get("vig_desde") or ""
    vig_hasta = filters.get("vig_hasta") or ""
    if vig_desde and vig_hasta:
        title = f"{title} | {vig_desde} a {vig_hasta}"

    ws_dashboard.merge_cells("A1:H1")
    ws_dashboard["A1"] = title
    ws_dashboard["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws_dashboard["A1"].fill = primary_fill
    ws_dashboard["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dashboard.row_dimensions[1].height = 24

    ws_dashboard["A2"] = "Filtros aplicados"
    ws_dashboard["A2"].font = bold_font
    ws_dashboard["B2"] = (
        f"CIA: {filters.get('cia') or 'Todas'} | "
        f"RAMO: {filters.get('ramo') or 'Todos'} | "
        f"SAGT: {filters.get('sub_agente') or 'Todos'} | "
        f"EJECUTIVO: {filters.get('ejecutivo') or 'Todos'} | "
        f"MONEDA: {filters.get('moneda') or 'Todas'}"
    )

    kpi_rows = [
        ("Cantidad de polizas", dashboard["totales"]["cantidad"]),
        ("Prima neta total", dashboard["totales"]["prima_neta"]),
        ("Prima total vendida", dashboard["totales"]["prima_total"]),
        ("Comision CIA total", dashboard["totales"]["comision_cia"]),
        ("Comision SAGT total", dashboard["totales"]["comision_sagt"]),
        ("Comision total", dashboard["totales"]["total_comision"]),
        ("Productividad total %", dashboard["totales"]["productividad_pct"]),
    ]
    best_month = dashboard["best_month"]
    best_month_label = "Sin datos"
    best_month_value = 0.0
    if best_month:
        best_month_label = f"{best_month['mes']} {best_month['anio']}"
        best_month_value = best_month["prima_total"]

    best_day = dashboard["best_day"]
    best_day_label = "Sin datos"
    best_day_value = 0.0
    if best_day:
        best_day_label = best_day["fecha"]
        best_day_value = best_day["prima_total"]

    dashboard_positions = [
        ("A4", kpi_rows[0][0], kpi_rows[0][1], False),
        ("C4", kpi_rows[1][0], kpi_rows[1][1], True),
        ("E4", kpi_rows[2][0], kpi_rows[2][1], True),
        ("A7", kpi_rows[3][0], kpi_rows[3][1], True),
        ("C7", kpi_rows[4][0], kpi_rows[4][1], True),
        ("E7", kpi_rows[5][0], kpi_rows[5][1], True),
        ("A10", kpi_rows[6][0], kpi_rows[6][1], False),
        ("C10", "Mes con mayor venta", best_month_value, True),
        ("E10", "Dia con mayor venta", best_day_value, True),
    ]

    for cell_ref, label, value, is_money in dashboard_positions:
        label_cell = ws_dashboard[cell_ref]
        value_cell = ws_dashboard.cell(row=label_cell.row, column=label_cell.column + 1)
        label_cell.value = label
        label_cell.fill = secondary_fill if label != "Mes con mayor venta" else warning_fill
        label_cell.font = bold_font
        label_cell.border = border
        value_cell.border = border
        value_cell.fill = accent_fill
        value_cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")
        if label == "Mes con mayor venta":
            value_cell.value = best_month_label
            ws_dashboard.cell(row=label_cell.row + 1, column=label_cell.column).value = "Prima total del mejor mes"
            ws_dashboard.cell(row=label_cell.row + 1, column=label_cell.column).fill = warning_fill
            ws_dashboard.cell(row=label_cell.row + 1, column=label_cell.column).font = bold_font
            ws_dashboard.cell(row=label_cell.row + 1, column=label_cell.column).border = border
            best_value_cell = ws_dashboard.cell(row=label_cell.row + 1, column=label_cell.column + 1)
            best_value_cell.value = best_month_value
            best_value_cell.number_format = money_format
            best_value_cell.fill = accent_fill
            best_value_cell.border = border
            best_value_cell.alignment = Alignment(horizontal="right")
        elif label == "Dia con mayor venta":
            value_cell.value = best_day_label
            ws_dashboard.cell(row=label_cell.row + 1, column=label_cell.column).value = "Prima total del mejor dia"
            ws_dashboard.cell(row=label_cell.row + 1, column=label_cell.column).fill = warning_fill
            ws_dashboard.cell(row=label_cell.row + 1, column=label_cell.column).font = bold_font
            ws_dashboard.cell(row=label_cell.row + 1, column=label_cell.column).border = border
            best_value_cell = ws_dashboard.cell(row=label_cell.row + 1, column=label_cell.column + 1)
            best_value_cell.value = best_day_value
            best_value_cell.number_format = money_format
            best_value_cell.fill = accent_fill
            best_value_cell.border = border
            best_value_cell.alignment = Alignment(horizontal="right")
        else:
            value_cell.value = value
            if is_money:
                value_cell.number_format = money_format
            if label == "Productividad total %":
                value_cell.number_format = "0.00"

    ws_dashboard["A14"] = "Top 10 Ramos por Prima Total"
    ws_dashboard["A14"].font = Font(bold=True, size=11)
    top_ramos_headers = ["RAMO", "PRIMA TOTAL"]
    for idx, header in enumerate(top_ramos_headers, start=1):
        ws_dashboard.cell(row=15, column=idx, value=header)
    style_table_header(ws_dashboard, 15, len(top_ramos_headers))

    top_ramos = dashboard["top_ramos"] or [{"ramo": "SIN DATOS", "prima_total": 0.0}]
    for row_idx, item in enumerate(top_ramos, start=16):
        ws_dashboard.cell(row=row_idx, column=1, value=item["ramo"]).border = border
        value_cell = ws_dashboard.cell(row=row_idx, column=2, value=item["prima_total"])
        value_cell.number_format = money_format
        value_cell.border = border
    add_excel_table(ws_dashboard, 15, 15 + len(top_ramos), 2, "TopRamosProduccion")

    ws_dashboard["D14"] = "Mejor Mes por Anio"
    ws_dashboard["D14"].font = Font(bold=True, size=11)
    best_year_headers = ["ANIO", "MES", "PRIMA TOTAL"]
    for idx, header in enumerate(best_year_headers, start=4):
        ws_dashboard.cell(row=15, column=idx, value=header)
    style_table_header(ws_dashboard, 15, 6)

    best_months_by_year = dashboard["best_months_by_year"] or [{"anio": "", "mes": "SIN DATOS", "prima_total": 0.0}]
    for row_idx, item in enumerate(best_months_by_year, start=16):
        for col_idx, key in enumerate(("anio", "mes", "prima_total"), start=4):
            cell = ws_dashboard.cell(row=row_idx, column=col_idx, value=item[key])
            cell.border = border
            if key == "prima_total":
                cell.number_format = money_format
                cell.alignment = Alignment(horizontal="right")

    ws_dashboard["A33"] = "Top 15 Contratantes"
    ws_dashboard["A33"].font = Font(bold=True, size=11)
    contratante_headers = ["CONTRATANTE", "PRIMA TOTAL", "COMISION TOTAL", "PRODUCTIVIDAD %", "CANT. RAMOS", "RAMOS"]
    for idx, header in enumerate(contratante_headers, start=1):
        ws_dashboard.cell(row=34, column=idx, value=header)
    style_table_header(ws_dashboard, 34, len(contratante_headers))

    top_contratantes = dashboard["top_contratantes"] or [
        {
            "contratante": "SIN DATOS",
            "prima_total": 0.0,
            "total_comision": 0.0,
            "productividad_pct": 0.0,
            "cantidad_ramos": 0,
            "ramos": "",
        }
    ]
    for row_idx, item in enumerate(top_contratantes[:10], start=35):
        ws_dashboard.cell(row=row_idx, column=1, value=item["contratante"]).border = border
        prima_cell = ws_dashboard.cell(row=row_idx, column=2, value=item["prima_total"])
        prima_cell.number_format = money_format
        prima_cell.border = border
        com_cell = ws_dashboard.cell(row=row_idx, column=3, value=item["total_comision"])
        com_cell.number_format = money_format
        com_cell.border = border
        prod_cell = ws_dashboard.cell(row=row_idx, column=4, value=item["productividad_pct"])
        prod_cell.number_format = "0.00"
        prod_cell.border = border
        cant_ramos_cell = ws_dashboard.cell(row=row_idx, column=5, value=item["cantidad_ramos"])
        cant_ramos_cell.border = border
        cant_ramos_cell.alignment = Alignment(horizontal="right")
        ws_dashboard.cell(row=row_idx, column=6, value=item["ramos"]).border = border

    ws_dashboard["H33"] = "Venta Baja / Normal"
    ws_dashboard["H33"].font = Font(bold=True, size=11)
    low_headers = ["CONTRATANTE", "PRIMA TOTAL", "CANT. RAMOS", "RAMOS"]
    for idx, header in enumerate(low_headers, start=8):
        ws_dashboard.cell(row=34, column=idx, value=header)
    for col in range(8, 12):
        cell = ws_dashboard.cell(row=34, column=col)
        cell.fill = primary_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    low_contratantes = dashboard["low_contratantes"] or [
        {
            "contratante": "SIN DATOS",
            "prima_total": 0.0,
            "cantidad_ramos": 0,
            "ramos": "",
        }
    ]
    for row_idx, item in enumerate(low_contratantes[:10], start=35):
        ws_dashboard.cell(row=row_idx, column=8, value=item["contratante"]).border = border
        low_prima_cell = ws_dashboard.cell(row=row_idx, column=9, value=item["prima_total"])
        low_prima_cell.number_format = money_format
        low_prima_cell.border = border
        low_cant_ramos_cell = ws_dashboard.cell(row=row_idx, column=10, value=item["cantidad_ramos"])
        low_cant_ramos_cell.border = border
        low_cant_ramos_cell.alignment = Alignment(horizontal="right")
        ws_dashboard.cell(row=row_idx, column=11, value=item["ramos"]).border = border

    daily_headers = [
        "FECHA",
        "ANIO",
        "MES NRO",
        "MES",
        "DIA",
        "CANTIDAD",
        "PRIMA NETA",
        "PRIMA TOTAL",
        "COMISION CIA",
        "COMISION SAGT",
        "COMISION TOTAL",
        "PRODUCTIVIDAD %",
    ]
    for idx, header in enumerate(daily_headers, start=1):
        ws_daily.cell(row=1, column=idx, value=header)
    style_table_header(ws_daily, 1, len(daily_headers))

    daily_records = dashboard["daily_records"] or [
        {
            "fecha": "SIN DATOS",
            "anio": "",
            "mes_numero": "",
            "mes": "",
            "dia": "",
            "cantidad": 0,
            "prima_neta": 0.0,
            "prima_total": 0.0,
            "comision_cia": 0.0,
            "comision_sagt": 0.0,
            "total_comision": 0.0,
            "productividad_pct": 0.0,
        }
    ]
    for row_idx, item in enumerate(daily_records, start=2):
        values = [
            item["fecha"],
            item["anio"],
            item["mes_numero"],
            item["mes"],
            item["dia"],
            item["cantidad"],
            item["prima_neta"],
            item["prima_total"],
            item["comision_cia"],
            item["comision_sagt"],
            item["total_comision"],
            item["productividad_pct"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws_daily.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if 7 <= col_idx <= 11:
                cell.number_format = money_format
            elif col_idx == 12:
                cell.number_format = "0.00"
            if col_idx >= 6:
                cell.alignment = Alignment(horizontal="right")
    add_excel_table(ws_daily, 1, 1 + len(daily_records), len(daily_headers), "ProductividadDiaria")
    ws_daily.freeze_panes = "A2"

    monthly_headers = [
        "PERIODO",
        "ANIO",
        "MES NRO",
        "MES",
        "CANTIDAD",
        "PRIMA NETA",
        "PRIMA TOTAL",
        "COMISION CIA",
        "COMISION SAGT",
        "COMISION TOTAL",
        "PRODUCTIVIDAD %",
    ]
    for idx, header in enumerate(monthly_headers, start=1):
        ws_monthly.cell(row=1, column=idx, value=header)
    style_table_header(ws_monthly, 1, len(monthly_headers))

    monthly_records = dashboard["monthly_records"] or [
        {
            "periodo": "SIN DATOS",
            "anio": "",
            "mes_numero": "",
            "mes": "",
            "cantidad": 0,
            "prima_neta": 0.0,
            "prima_total": 0.0,
            "comision_cia": 0.0,
            "comision_sagt": 0.0,
        }
    ]
    for row_idx, item in enumerate(monthly_records, start=2):
        values = [
            item["periodo"],
            item["anio"],
            item["mes_numero"],
            item["mes"],
            item["cantidad"],
            item["prima_neta"],
            item["prima_total"],
            item["comision_cia"],
            item["comision_sagt"],
            item["total_comision"],
            item["productividad_pct"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws_monthly.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if 6 <= col_idx <= 10:
                cell.number_format = money_format
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 11:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right")
    add_excel_table(ws_monthly, 1, 1 + len(monthly_records), len(monthly_headers), "ResumenMensualProduccion")
    ws_monthly.freeze_panes = "A2"

    yearly_headers = [
        "ANIO",
        "ENERO",
        "FEBRERO",
        "MARZO",
        "ABRIL",
        "MAYO",
        "JUNIO",
        "JULIO",
        "AGOSTO",
        "SETIEMBRE",
        "OCTUBRE",
        "NOVIEMBRE",
        "DICIEMBRE",
        "TOTAL PRIMA NETA",
        "TOTAL PRIMA TOTAL",
        "TOTAL COMISION CIA",
        "TOTAL COMISION SAGT",
        "TOTAL COMISION",
        "PRODUCTIVIDAD %",
        "CANTIDAD",
    ]
    for idx, header in enumerate(yearly_headers, start=1):
        ws_yearly.cell(row=1, column=idx, value=header)
    style_table_header(ws_yearly, 1, len(yearly_headers))

    yearly_records = dashboard["yearly_records"] or [
        {
            "anio": "SIN DATOS",
            "meses": [0.0] * 12,
            "prima_neta": 0.0,
            "prima_total": 0.0,
            "comision_cia": 0.0,
            "comision_sagt": 0.0,
            "cantidad": 0,
        }
    ]
    for row_idx, item in enumerate(yearly_records, start=2):
        values = [
            item["anio"],
            *item["meses"],
            item["prima_neta"],
            item["prima_total"],
            item["comision_cia"],
            item["comision_sagt"],
            item["total_comision"],
            item["productividad_pct"],
            item["cantidad"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws_yearly.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if 2 <= col_idx <= 18:
                cell.number_format = money_format
            elif col_idx == 19:
                cell.number_format = "0.00"
            if col_idx != 1:
                cell.alignment = Alignment(horizontal="right")
    add_excel_table(ws_yearly, 1, 1 + len(yearly_records), len(yearly_headers), "ResumenAnualProduccion")
    ws_yearly.freeze_panes = "A2"
    if len(yearly_records) >= 1:
        ws_yearly.conditional_formatting.add(
            f"B2:M{1 + len(yearly_records)}",
            ColorScaleRule(
                start_type="min",
                start_color="FFF2CC",
                mid_type="percentile",
                mid_value=50,
                mid_color="9FD5B3",
                end_type="max",
                end_color="00B050",
            ),
        )

    render_ramo_matrix(ws_ramo_matrix)
    render_contratante_matrix(ws_contratante_matrix)

    ramo_headers = [
        "RAMO",
        "CANTIDAD",
        "PRIMA NETA",
        "PRIMA TOTAL",
        "COMISION CIA",
        "COMISION SAGT",
        "COMISION TOTAL",
        "PRODUCTIVIDAD %",
    ]
    for idx, header in enumerate(ramo_headers, start=1):
        ws_ramo.cell(row=1, column=idx, value=header)
    style_table_header(ws_ramo, 1, len(ramo_headers))

    ramo_records = dashboard["ramo_records"] or [
        {
            "ramo": "SIN DATOS",
            "cantidad": 0,
            "prima_neta": 0.0,
            "prima_total": 0.0,
            "comision_cia": 0.0,
            "comision_sagt": 0.0,
            "total_comision": 0.0,
            "productividad_pct": 0.0,
        }
    ]
    for row_idx, item in enumerate(ramo_records, start=2):
        values = [
            item["ramo"],
            item["cantidad"],
            item["prima_neta"],
            item["prima_total"],
            item["comision_cia"],
            item["comision_sagt"],
            item["total_comision"],
            item["productividad_pct"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws_ramo.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if 3 <= col_idx <= 7:
                cell.number_format = money_format
            elif col_idx == 8:
                cell.number_format = "0.00"
            if col_idx >= 2:
                cell.alignment = Alignment(horizontal="right")
    add_excel_table(ws_ramo, 1, 1 + len(ramo_records), len(ramo_headers), "ResumenRamos")
    ws_ramo.freeze_panes = "A2"

    contratante_headers = [
        "CONTRATANTE",
        "CANTIDAD",
        "CANT. RAMOS",
        "RAMOS",
        "PRIMA NETA",
        "PRIMA TOTAL",
        "COMISION CIA",
        "COMISION SAGT",
        "COMISION TOTAL",
        "PRODUCTIVIDAD %",
    ]
    for idx, header in enumerate(contratante_headers, start=1):
        ws_contratante.cell(row=1, column=idx, value=header)
    style_table_header(ws_contratante, 1, len(contratante_headers))

    contratante_records = dashboard["contratante_records"] or [
        {
            "contratante": "SIN DATOS",
            "cantidad": 0,
            "prima_neta": 0.0,
            "prima_total": 0.0,
            "comision_cia": 0.0,
            "comision_sagt": 0.0,
            "total_comision": 0.0,
            "productividad_pct": 0.0,
        }
    ]
    for row_idx, item in enumerate(contratante_records, start=2):
        values = [
            item["contratante"],
            item["cantidad"],
            item["cantidad_ramos"],
            item["ramos"],
            item["prima_neta"],
            item["prima_total"],
            item["comision_cia"],
            item["comision_sagt"],
            item["total_comision"],
            item["productividad_pct"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws_contratante.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if 5 <= col_idx <= 9:
                cell.number_format = money_format
            elif col_idx == 10:
                cell.number_format = "0.00"
            if col_idx in {2, 3, 5, 6, 7, 8, 9, 10}:
                cell.alignment = Alignment(horizontal="right")
    add_excel_table(ws_contratante, 1, 1 + len(contratante_records), len(contratante_headers), "ResumenContratantes")
    ws_contratante.freeze_panes = "A2"

    productivity_headers = [
        "ANIO",
        "PRIMA NETA",
        "PRIMA TOTAL",
        "COMISION CIA",
        "COMISION SAGT",
        "COMISION TOTAL",
        "PRODUCTIVIDAD %",
        "CANTIDAD",
    ]
    for idx, header in enumerate(productivity_headers, start=1):
        ws_productivity.cell(row=1, column=idx, value=header)
    style_table_header(ws_productivity, 1, len(productivity_headers))

    productivity_records = dashboard["yearly_records"] or [
        {
            "anio": "SIN DATOS",
            "prima_neta": 0.0,
            "prima_total": 0.0,
            "comision_cia": 0.0,
            "comision_sagt": 0.0,
            "total_comision": 0.0,
            "productividad_pct": 0.0,
            "cantidad": 0,
        }
    ]
    for row_idx, item in enumerate(productivity_records, start=2):
        values = [
            item["anio"],
            item["prima_neta"],
            item["prima_total"],
            item["comision_cia"],
            item["comision_sagt"],
            item["total_comision"],
            item["productividad_pct"],
            item["cantidad"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws_productivity.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if 2 <= col_idx <= 6:
                cell.number_format = money_format
            elif col_idx == 7:
                cell.number_format = "0.00"
            if col_idx >= 2:
                cell.alignment = Alignment(horizontal="right")
    add_excel_table(ws_productivity, 1, 1 + len(productivity_records), len(productivity_headers), "ProductividadAnual")
    ws_productivity.freeze_panes = "A2"

    detail_title = "Detalle de Produccion"
    ws_detail.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws_detail["A1"] = detail_title
    ws_detail["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws_detail["A1"].fill = primary_fill
    ws_detail["A1"].alignment = Alignment(horizontal="center")

    for col, header in enumerate(headers, start=1):
        ws_detail.cell(row=2, column=col, value=header)
    style_table_header(ws_detail, 2, len(headers))

    money_cols = {15, 16, 18, 21}
    percent_cols = {17, 20}
    for row_idx, row in enumerate(table_rows, start=3):
        fill = PatternFill("solid", fgColor="F8FBFF") if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, value in enumerate(row, start=1):
            if col_idx == 1:
                continue
            excel_col_idx = col_idx - 1
            cell = ws_detail.cell(row=row_idx, column=excel_col_idx, value=value)
            cell.border = border
            cell.fill = fill
            cell.font = Font(size=9)
            if excel_col_idx in money_cols:
                cell.number_format = money_format
                cell.alignment = Alignment(horizontal="right")
            elif excel_col_idx in percent_cols:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
            if excel_col_idx == 8 and rows[row_idx - 3].get("es_financiamiento_grupal"):
                cell.font = Font(size=9, bold=True, color="7A3DB8")

    detail_total_row = len(table_rows) + 3
    ws_detail.cell(row=detail_total_row, column=1, value="TOTAL").font = bold_font
    for col in [15, 16, 18, 21]:
        cell = ws_detail.cell(row=detail_total_row, column=col)
        cell.font = bold_font
        cell.number_format = money_format
        cell.alignment = Alignment(horizontal="right")
        if table_rows:
            col_letter = get_column_letter(col)
            cell.value = f"=SUBTOTAL(109,{col_letter}3:{col_letter}{detail_total_row - 1})"
        else:
            cell.value = 0

    detail_col_widths = [
        12, 28, 28, 28, 16, 18, 18, 16, 8, 16,
        16, 12, 12, 10, 14, 14, 10, 14, 16, 10,
        14, 14, 18, 18, 14, 16, 16, 18, 16, 12,
    ]
    for idx, width in enumerate(detail_col_widths, start=1):
        ws_detail.column_dimensions[get_column_letter(idx)].width = width
    add_excel_table(ws_detail, 2, max(detail_total_row - 1, 2), len(headers), "DetalleProduccion")
    ws_detail.freeze_panes = "A3"
    ws_detail.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{max(detail_total_row - 1, 2)}"

    dashboard_chart = BarChart()
    dashboard_chart.type = "col"
    dashboard_chart.style = 10
    dashboard_chart.title = "Prima Total por Mes"
    dashboard_chart.y_axis.title = "Prima Total"
    dashboard_chart.x_axis.title = "Periodo"
    if dashboard["monthly_records"]:
        data = Reference(ws_monthly, min_col=7, min_row=1, max_row=1 + len(dashboard["monthly_records"]))
        categories = Reference(ws_monthly, min_col=1, min_row=2, max_row=1 + len(dashboard["monthly_records"]))
        dashboard_chart.add_data(data, titles_from_data=True)
        dashboard_chart.set_categories(categories)
        dashboard_chart.height = 8
        dashboard_chart.width = 18
        ws_dashboard.add_chart(dashboard_chart, "H3")

    for sheet in (ws_dashboard, ws_daily, ws_monthly, ws_yearly, ws_ramo_matrix, ws_contratante_matrix, ws_ramo, ws_contratante, ws_productivity):
        for col_idx in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 16
    ws_ramo_matrix.column_dimensions["A"].width = 32
    ws_contratante_matrix.column_dimensions["A"].width = 42
    ws_contratante_matrix.column_dimensions["B"].width = 32
    ws_ramo.column_dimensions["A"].width = 32
    ws_contratante.column_dimensions["A"].width = 42
    ws_contratante.column_dimensions["D"].width = 46
    ws_dashboard.column_dimensions["A"].width = 42
    ws_dashboard.column_dimensions["B"].width = 16
    ws_dashboard.column_dimensions["C"].width = 22
    ws_dashboard.column_dimensions["D"].width = 16
    ws_dashboard.column_dimensions["F"].width = 44
    ws_dashboard.column_dimensions["H"].width = 34
    ws_dashboard.column_dimensions["I"].width = 16
    ws_dashboard.column_dimensions["K"].width = 44

    wb.save(filepath)
    return filepath, filename
