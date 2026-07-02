from datetime import datetime

from models.db import get_connection


POLIZA_EXPR = """
COALESCE(
    CAST(AES_DECRYPT(FROM_BASE64(p.poliza), @SIS_KEY) AS CHAR),
    CAST(AES_DECRYPT(p.poliza, @SIS_KEY) AS CHAR),
    p.poliza
)
"""

RECIBO_EXPR = """
COALESCE(
    CAST(AES_DECRYPT(FROM_BASE64(p.recibo), @SIS_KEY) AS CHAR),
    CAST(AES_DECRYPT(p.recibo, @SIS_KEY) AS CHAR),
    p.recibo
)
"""

PLANILLA_EXPR = """
COALESCE(
    CAST(AES_DECRYPT(FROM_BASE64(p.contrato_nro), @SIS_KEY) AS CHAR),
    CAST(AES_DECRYPT(p.contrato_nro, @SIS_KEY) AS CHAR),
    p.contrato_nro
)
"""

CERTIF_EXPR = """
COALESCE(
    CAST(AES_DECRYPT(FROM_BASE64(p.nro), @SIS_KEY) AS CHAR),
    CAST(AES_DECRYPT(p.nro, @SIS_KEY) AS CHAR),
    p.nro
)
"""

CONTRATANTE_EXPR = """
COALESCE(
    CAST(AES_DECRYPT(FROM_BASE64(c.razon_social), @SIS_KEY) AS CHAR),
    CAST(AES_DECRYPT(c.razon_social, @SIS_KEY) AS CHAR),
    c.razon_social
)
"""

ASEGURADO_EXPR = """
COALESCE(
    CAST(AES_DECRYPT(FROM_BASE64(p.asegurado), @SIS_KEY) AS CHAR),
    CAST(AES_DECRYPT(p.asegurado, @SIS_KEY) AS CHAR),
    p.asegurado
)
"""

CONTRATANTE_REPORTE_EXPR = f"""
COALESCE(
    {CONTRATANTE_EXPR},
    {ASEGURADO_EXPR}
)
"""

USO_EXPR = "JSON_UNQUOTE(JSON_EXTRACT(p.datos_vehiculo, '$.uso'))"
CLASE_EXPR = "JSON_UNQUOTE(JSON_EXTRACT(p.datos_vehiculo, '$.clase'))"
PLACA_EXPR = "JSON_UNQUOTE(JSON_EXTRACT(p.datos_vehiculo, '$.placa'))"
PROD_ARIAS_EXPR = "ROUND(COALESCE(p.imp_compania, 0) - COALESCE(p.imp_subagente, 0), 2)"


def _build_where(search='', fecha_desde=None, fecha_hasta=None):
    where_clauses = [
        "(p.ramo LIKE '%SOAT%' OR COALESCE(p.ramos_producto, '') LIKE '%SOAT%')",
        "COALESCE(p.activo, 1) = 1",
    ]
    params = []

    if search:
        where_clauses.append(
            "("
            f"{POLIZA_EXPR} COLLATE utf8mb4_0900_ai_ci LIKE %s "
            f"OR {RECIBO_EXPR} COLLATE utf8mb4_0900_ai_ci LIKE %s "
            f"OR {PLANILLA_EXPR} COLLATE utf8mb4_0900_ai_ci LIKE %s "
            f"OR {CERTIF_EXPR} COLLATE utf8mb4_0900_ai_ci LIKE %s "
            f"OR {PLACA_EXPR} COLLATE utf8mb4_0900_ai_ci LIKE %s "
            f"OR {USO_EXPR} COLLATE utf8mb4_0900_ai_ci LIKE %s "
            f"OR {CLASE_EXPR} COLLATE utf8mb4_0900_ai_ci LIKE %s "
            f"OR {CONTRATANTE_EXPR} COLLATE utf8mb4_0900_ai_ci LIKE %s "
            "OR p.codigo_agente COLLATE utf8mb4_0900_ai_ci LIKE %s "
            "OR p.ejecutivo COLLATE utf8mb4_0900_ai_ci LIKE %s"
            ")"
        )
        s = f"%{search}%"
        params += [s] * 10

    if fecha_desde:
        where_clauses.append("p.fecha_emision >= %s")
        params.append(fecha_desde)

    if fecha_hasta:
        where_clauses.append("p.fecha_emision <= %s")
        params.append(fecha_hasta)

    return " AND ".join(where_clauses), params


def _date_label(value):
    if not value:
        return ''
    try:
        return datetime.strptime(str(value), '%Y-%m-%d').strftime('%d/%m/%Y')
    except Exception:
        return str(value)


def _get_vendedor_labels(rows):
    vendedores = sorted({
        (row.get('vendedor') or '').strip().upper()
        for row in (rows or [])
        if (row.get('vendedor') or '').strip()
    })

    if len(vendedores) == 1:
        vendedor = vendedores[0]
        if ',' in vendedor:
            corto = vendedor.split(',')[-1].strip()
        else:
            partes = [p for p in vendedor.split() if p]
            corto = partes[-1] if partes else vendedor
        return vendedor, corto or 'MONICA'

    if len(vendedores) > 1:
        return 'COMERCIALIZADORES', 'COMERCIAL'

    return 'COMERCIALIZADOR', 'MONICA'


def _get_reporte_rows(cur, where_sql, params, per_page=None, offset=0):
    query = f"""
        SELECT
            p.idPoliza,
            {POLIZA_EXPR} AS poliza,
            {RECIBO_EXPR} AS recibo,
            {PLANILLA_EXPR} AS planilla,
            {CERTIF_EXPR} AS certif,
            p.codigo_agente AS codigo,
            p.ejecutivo AS vendedor,
            p.moneda,
            ROUND(COALESCE(p.prima_comercial_igv, 0), 2) AS costo,
            ROUND(COALESCE(p.prima_neta, 0), 2) AS prima_neta,
            ROUND(COALESCE(p.prima_comercial_igv, 0), 2) AS prima_comercial_igv,
            ROUND(COALESCE(p.porc_compania, 0), 2) AS pct_total,
            ROUND(COALESCE(p.imp_compania, 0), 2) AS prod_total,
            ROUND(COALESCE(p.porc_subagente, 0), 2) AS pct_comerc,
            ROUND(COALESCE(p.imp_subagente, 0), 2) AS prod_monica,
            {PROD_ARIAS_EXPR} AS prod_arias,
            p.vig_desde,
            p.vig_hasta,
            p.fecha_emision,
            p.cia,
            {USO_EXPR} AS tipo,
            {CLASE_EXPR} AS tipo_vehiculo,
            {PLACA_EXPR} AS placa,
            {CONTRATANTE_REPORTE_EXPR} AS contratante,
            {ASEGURADO_EXPR} AS asegurado,
            p.estado
        FROM polizas p
        LEFT JOIN clientes c ON c.idCliente = p.cliente_id
        WHERE {where_sql}
        ORDER BY p.fecha_emision DESC, p.creado_en DESC
    """
    query_params = list(params)

    if per_page and per_page > 0:
        query += "\n LIMIT %s OFFSET %s"
        query_params += [per_page, offset]

    cur.execute(query, query_params)
    return cur.fetchall()


def get_produccion_soat(page=1, per_page=20, search='', fecha_desde=None, fecha_hasta=None):
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        safe_page = max(int(page or 1), 1)
        safe_per_page = int(per_page or 20)
        offset = (safe_page - 1) * safe_per_page if safe_per_page > 0 else 0
        where_sql, params = _build_where(search=search, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)

        cur.execute(
            f"""
            SELECT COUNT(*) AS total
            FROM polizas p
            LEFT JOIN clientes c ON c.idCliente = p.cliente_id
            WHERE {where_sql}
            """,
            params,
        )
        total = cur.fetchone()['total']

        rows = _get_reporte_rows(cur, where_sql, params, per_page=safe_per_page, offset=offset)

        cur.execute(
            f"""
            SELECT
                ROUND(SUM(COALESCE(p.prima_comercial_igv, 0)), 2) AS total_costo,
                ROUND(SUM(COALESCE(p.prima_neta, 0)), 2) AS total_prima_neta,
                ROUND(SUM(COALESCE(p.imp_compania, 0)), 2) AS total_prod_total,
                ROUND(SUM(COALESCE(p.imp_subagente, 0)), 2) AS total_prod_monica,
                ROUND(SUM(COALESCE(p.imp_compania, 0) - COALESCE(p.imp_subagente, 0)), 2) AS total_prod_arias
            FROM polizas p
            LEFT JOIN clientes c ON c.idCliente = p.cliente_id
            WHERE {where_sql}
            """,
            params,
        )
        totales = cur.fetchone()

        return {'rows': rows, 'total': total, 'totales': totales}
    finally:
        cur.close()
        conn.close()


def export_produccion_soat_excel(search='', fecha_desde=None, fecha_hasta=None):
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        where_sql, params = _build_where(search=search, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
        rows = _get_reporte_rows(cur, where_sql, params)
    finally:
        cur.close()
        conn.close()

    # Generate Excel
    import os
    from flask import current_app
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    headers = [
        "CIA", "TIPO", "CERTIF.", "Mo", "COSTO", "P. NETA", "PLACA", "VIG.INI",
        "POLIZA", "CONTRATANTE", "TIPO DE VEHICULOS", "% TOTAL", "PROD.TOTAL",
        "% COMERC.", "", "PROD. ARIAS"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Produccion SOAT"
    bold_font = Font(bold=True)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    title_fill = PatternFill('solid', fgColor='D9D9D9')
    monica_fill = PatternFill('solid', fgColor='92D050')
    pagos_fill = PatternFill('solid', fgColor='5B9BD5')
    comercial_fill = PatternFill('solid', fgColor='92D050')
    arias_fill = PatternFill('solid', fgColor='D9B8D8')
    header_fill = PatternFill('solid', fgColor='F8CBAD')

    vendedor_label, vendedor_corto = _get_vendedor_labels(rows)
    headers[14] = f"PROD. {vendedor_corto}"

    desde_label = _date_label(fecha_desde) or "INICIO"
    hasta_label = _date_label(fecha_hasta) or "FIN"
    debug_search = search or "(vacio)"
    debug_desde = _date_label(fecha_desde) or "(vacio)"
    debug_hasta = _date_label(fecha_hasta) or "(vacio)"
    debug_text = (
        f"DEBUG EXPORTACION | registros: {len(rows)} | "
        f"search: {debug_search} | fecha_desde: {debug_desde} | fecha_hasta: {debug_hasta}"
    )

    debug_fill = PatternFill('solid', fgColor='FFF2CC')

    ws.merge_cells('A1:P1')
    ws['A1'] = debug_text
    ws['A1'].font = Font(bold=True, size=10)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A1'].fill = debug_fill
    ws['A1'].border = border

    ws.merge_cells('A2:K2')
    ws['A2'] = f'Ventas realizadas del {desde_label} al {hasta_label}'
    ws['A2'].font = Font(bold=True, size=16)
    ws['A2'].alignment = center
    ws['A2'].fill = title_fill

    ws.merge_cells('L2:M2')
    ws['L2'] = 'PAGOS GENERALES'
    ws['L2'].font = bold_font
    ws['L2'].alignment = center
    ws['L2'].fill = pagos_fill

    ws.merge_cells('N2:O2')
    ws['N2'] = 'PAGO A COMERCIALIZADOR'
    ws['N2'].font = bold_font
    ws['N2'].alignment = center
    ws['N2'].fill = comercial_fill

    ws.merge_cells('P2:P3')
    ws['P2'] = 'ARIAS'
    ws['P2'].font = bold_font
    ws['P2'].alignment = center
    ws['P2'].fill = arias_fill

    ws.merge_cells('A3:K3')
    ws['A3'] = vendedor_label
    ws['A3'].font = Font(bold=True, size=14)
    ws['A3'].alignment = center
    ws['A3'].fill = monica_fill

    for row_num in (2, 3):
        for col in range(1, 17):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = bold_font
        cell.alignment = center
        cell.border = border
        if col in (12, 13):
            cell.fill = pagos_fill
        elif col in (14, 15):
            cell.fill = comercial_fill
        elif col == 16:
            cell.fill = arias_fill
        else:
            cell.fill = header_fill

    currency_format = '"S/" #,##0.00'
    percent_format = '0.00'
    date_format = 'DD/MM/YYYY'

    widths = {
        'A': 10, 'B': 14, 'C': 12, 'D': 7, 'E': 12, 'F': 12, 'G': 12, 'H': 12,
        'I': 14, 'J': 34, 'K': 20, 'L': 10, 'M': 12, 'N': 10, 'O': 14, 'P': 14
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    for r_idx, row in enumerate(rows, start=5):
        ws.cell(row=r_idx, column=1, value=row.get('cia'))
        ws.cell(row=r_idx, column=2, value=row.get('tipo'))
        ws.cell(row=r_idx, column=3, value=row.get('certif'))
        ws.cell(row=r_idx, column=4, value=row.get('moneda'))

        c = ws.cell(row=r_idx, column=5, value=float(row.get('costo') or 0))
        c.number_format = currency_format

        c = ws.cell(row=r_idx, column=6, value=float(row.get('prima_neta') or 0))
        c.number_format = currency_format

        ws.cell(row=r_idx, column=7, value=row.get('placa'))

        c = ws.cell(row=r_idx, column=8, value=row.get('vig_desde'))
        c.number_format = date_format

        ws.cell(row=r_idx, column=9, value=row.get('poliza'))
        ws.cell(row=r_idx, column=10, value=row.get('contratante'))
        ws.cell(row=r_idx, column=11, value=row.get('tipo_vehiculo'))

        c = ws.cell(row=r_idx, column=12, value=float(row.get('pct_total') or 0))
        c.number_format = percent_format

        c = ws.cell(row=r_idx, column=13, value=float(row.get('prod_total') or 0))
        c.number_format = currency_format

        c = ws.cell(row=r_idx, column=14, value=float(row.get('pct_comerc') or 0))
        c.number_format = percent_format

        c = ws.cell(row=r_idx, column=15, value=float(row.get('prod_monica') or 0))
        c.number_format = currency_format

        c = ws.cell(row=r_idx, column=16, value=float(row.get('prod_arias') or 0))
        c.number_format = currency_format

        for col in range(1, 17):
            cell = ws.cell(row=r_idx, column=col)
            cell.border = border

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:P{max(5, len(rows) + 4)}'

    # Save file
    try:
        upload_folder = current_app.config.get("UPLOAD_FOLDER")
    except Exception:
        upload_folder = None
        
    if not upload_folder:
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
        upload_folder = os.path.join(base_dir, "uploads")

    exports_dir = os.path.join(upload_folder, "exports")
    os.makedirs(exports_dir, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"produccion_soat_{ts}.xlsx"
    filepath = os.path.join(exports_dir, filename)
    
    wb.save(filepath)
    return filepath, filename



