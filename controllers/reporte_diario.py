from models.db import get_connection
from datetime import date, datetime, timedelta
from decimal import Decimal
import os


def get_reporte_diario_data(filters=None):
    """Retorna pólizas creadas hoy o filtradas por usuario/fecha de registro (gestión diaria)."""
    conn = get_connection()
    rows = []
    try:
        cur = conn.cursor(dictionary=True)
        filters = filters or {}
        ref_date = (datetime.utcnow() - timedelta(hours=12)).date().isoformat()
        f_reg_desde = (filters.get("f_reg_desde") or "").strip()
        f_reg_hasta = (filters.get("f_reg_hasta") or "").strip()
        usuario = (filters.get("usuario") or "").strip()

        where = ["p.activo = 1", "(p.anulado = 0 OR p.anulado IS NULL)"]
        params = []

        created_day_expr = "DATE(p.creado_en - INTERVAL 7 HOUR)"

        if f_reg_desde and f_reg_hasta and f_reg_desde == f_reg_hasta:
            where.append(f"{created_day_expr} = %s")
            params.append(f_reg_desde)
        else:
            if f_reg_desde:
                where.append(f"{created_day_expr} >= %s")
                params.append(f_reg_desde)
            if f_reg_hasta:
                where.append(f"{created_day_expr} <= %s")
                params.append(f_reg_hasta)

        if not f_reg_desde and not f_reg_hasta:
            where.append(f"{created_day_expr} = %s")
            params.append(ref_date)

        if usuario:
            where.append(
                "("
                "TRIM(p.usuario_registro) COLLATE utf8mb4_unicode_ci = TRIM(%s) COLLATE utf8mb4_unicode_ci "
                "OR TRIM(p.usuario_registro) COLLATE utf8mb4_unicode_ci = ("
                "    SELECT TRIM(username) COLLATE utf8mb4_unicode_ci "
                "    FROM usuarios "
                "    WHERE TRIM(COALESCE(NULLIF(nombre, ''), username)) COLLATE utf8mb4_unicode_ci = TRIM(%s) COLLATE utf8mb4_unicode_ci "
                "    LIMIT 1"
                ") "
                "OR TRIM(p.usuario_registro) COLLATE utf8mb4_unicode_ci = ("
                "    SELECT TRIM(COALESCE(NULLIF(nombre, ''), username)) COLLATE utf8mb4_unicode_ci "
                "    FROM usuarios "
                "    WHERE TRIM(username) COLLATE utf8mb4_unicode_ci = TRIM(%s) COLLATE utf8mb4_unicode_ci "
                "    LIMIT 1"
                ")"
                ")"
            )
            params.extend([usuario, usuario, usuario])

        where_clause = " WHERE " + " AND ".join(where)
        sql = """
            SELECT
                p.idPoliza,
                COALESCE(
                    CAST(AES_DECRYPT(FROM_BASE64(p.poliza), @SIS_KEY) AS CHAR),
                    CAST(AES_DECRYPT(p.poliza, @SIS_KEY) AS CHAR),
                    p.poliza
                ) AS poliza,
                COALESCE(
                    CAST(AES_DECRYPT(FROM_BASE64(p.recibo), @SIS_KEY) AS CHAR),
                    CAST(AES_DECRYPT(p.recibo, @SIS_KEY) AS CHAR),
                    p.recibo
                ) AS recibo,
                COALESCE(
                    CAST(AES_DECRYPT(FROM_BASE64(p.contrato_nro), @SIS_KEY) AS CHAR),
                    CAST(AES_DECRYPT(p.contrato_nro, @SIS_KEY) AS CHAR),
                    p.contrato_nro
                ) AS contrato_nro,
                COALESCE(
                    CAST(AES_DECRYPT(FROM_BASE64(p.nro), @SIS_KEY) AS CHAR),
                    CAST(AES_DECRYPT(p.nro, @SIS_KEY) AS CHAR),
                    p.nro
                ) AS nro,
                COALESCE(
                    COALESCE(
                        CAST(AES_DECRYPT(FROM_BASE64(c.razon_social), @SIS_KEY) AS CHAR),
                        CAST(AES_DECRYPT(c.razon_social, @SIS_KEY) AS CHAR),
                        c.razon_social
                    ),
                    COALESCE(
                        CAST(AES_DECRYPT(FROM_BASE64(p.asegurado), @SIS_KEY) AS CHAR),
                        CAST(AES_DECRYPT(p.asegurado, @SIS_KEY) AS CHAR),
                        p.asegurado
                    )
                ) AS cliente,
                p.cia,
                p.ramo,
                p.ramos_producto,
                p.moneda,
                COALESCE(p.prima_comercial_igv) AS prima_total,
                p.vig_desde,
                p.vig_hasta,
                p.ejecutivo,
                p.sub_agente,
                p.estado,
                p.usuario_registro,
                p.creado_en
            FROM polizas p
            LEFT JOIN clientes c ON c.idCliente = p.cliente_id
            {where_clause}
            ORDER BY p.idPoliza DESC
        """
        sql = sql.format(where_clause=where_clause)
        cur.execute(sql, tuple(params))
        for row in cur.fetchall():
            for k, v in row.items():
                if isinstance(v, Decimal):
                    row[k] = float(v)
                elif hasattr(v, 'isoformat'):
                    row[k] = v.isoformat()
                elif v is None:
                    row[k] = None   # dejar None, jsonify lo convierte a null
            rows.append(row)
        cur.close()
    finally:
        conn.close()
    return rows


# ─────────────────────────────────────────────
#  Exportación
# ─────────────────────────────────────────────

HEADERS = [
    "N°", "Póliza", "Recibo", "Cliente", "Compañía", "Ramo", "Producto",
    "Moneda", "Prima Total", "Vig. Desde", "Vig. Hasta",
    "Ejecutivo", "Sub Agente", "Estado", "Registrado por", "Fecha/Hora registro",
]

def _build_table_rows(rows):
    result = []
    for i, r in enumerate(rows, 1):
        created = r.get("creado_en")
        created_str = ""
        if created:
            t = str(created).replace("T", " ")
            created_str = t[:19] if len(t) >= 19 else t
        result.append([
            i,
            r.get("poliza") or r.get("contrato_nro") or r.get("nro") or "",
            r.get("recibo") or "",
            r.get("cliente") or "",
            r.get("cia") or "",
            r.get("ramo") or "",
            r.get("ramos_producto") or "",
            r.get("moneda") or "",
            float(r.get("prima_total") or 0),
            r.get("vig_desde") or "",
            r.get("vig_hasta") or "",
            r.get("ejecutivo") or "",
            r.get("sub_agente") or "",
            r.get("estado") or "",
            r.get("usuario_registro") or "",
            created_str,
        ])
    return result


def export_excel(upload_folder: str, filters=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    rows = get_reporte_diario_data(filters)
    table_rows = _build_table_rows(rows)
    filters = filters or {}
    f_reg_desde = (filters.get("f_reg_desde") or "").strip()
    f_reg_hasta = (filters.get("f_reg_hasta") or "").strip()
    if f_reg_desde and f_reg_hasta and f_reg_desde != f_reg_hasta:
        today_str = f"{f_reg_desde} a {f_reg_hasta}"
    else:
        ref = f_reg_desde or f_reg_hasta or (datetime.utcnow() - timedelta(hours=12)).date().isoformat()
        today_str = datetime.fromisoformat(ref).strftime("%d/%m/%Y")

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Diario"

    # Título
    ws.merge_cells("A1:P1")
    title_cell = ws["A1"]
    title_cell.value = f"REPORTE DIARIO DE PÓLIZAS — {today_str}"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F59A3")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Cabecera
    header_fill = PatternFill("solid", fgColor="399AD6")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[2].height = 20

    # Datos
    for ri, row in enumerate(table_rows, 3):
        fill = PatternFill("solid", fgColor="F4F8FF") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.fill = fill
            cell.font = Font(size=9)
            if ci == 9:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif ci == 1:
                cell.alignment = Alignment(horizontal="center")

    # Fila totales (SUBTOTAL para que el total se actualice al filtrar en Excel)
    total_row = len(table_rows) + 3
    ws.cell(row=total_row, column=8, value="TOTAL").font = Font(bold=True, size=9)
    total_cell = ws.cell(row=total_row, column=9)
    total_cell.font = Font(bold=True, size=9)
    total_cell.number_format = '#,##0.00'
    total_cell.alignment = Alignment(horizontal="right")
    if table_rows:
        total_cell.value = f"=SUBTOTAL(109,I3:I{total_row-1})"
    else:
        total_cell.value = 0

    # Ancho columnas
    col_widths = [5, 16, 16, 30, 16, 22, 22, 10, 14, 13, 13, 16, 16, 12, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    exports_dir = os.path.join(upload_folder, "exports")
    os.makedirs(exports_dir, exist_ok=True)
    date_label = (datetime.utcnow() - timedelta(hours=12)).date().strftime("%d-%m-%Y")
    filename = f"Reporte Diario {date_label}.xlsx"
    filepath = os.path.join(exports_dir, filename)
    wb.save(filepath)
    return filepath, filename


def export_pdf(upload_folder: str, filters=None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from datetime import datetime

    rows = get_reporte_diario_data(filters)
    table_rows = _build_table_rows(rows)
    filters = filters or {}
    f_reg_desde = (filters.get("f_reg_desde") or "").strip()
    f_reg_hasta = (filters.get("f_reg_hasta") or "").strip()
    if f_reg_desde and f_reg_hasta and f_reg_desde != f_reg_hasta:
        today_str = f"{f_reg_desde} a {f_reg_hasta}"
    else:
        ref = f_reg_desde or f_reg_hasta or (datetime.utcnow() - timedelta(hours=12)).date().isoformat()
        today_str = datetime.fromisoformat(ref).strftime("%d/%m/%Y")

    exports_dir = os.path.join(upload_folder, "exports")
    os.makedirs(exports_dir, exist_ok=True)
    date_label = (datetime.utcnow() - timedelta(hours=12)).date().strftime("%d-%m-%Y")
    filename = f"Reporte Diario {date_label}.pdf"
    filepath = os.path.join(exports_dir, filename)

    # A4 landscape útil: 29.7cm - 1.6cm márgenes = 28.1 cm
    PAGE = landscape(A4)
    LEFT = RIGHT = 0.8 * cm
    TOP = BOTTOM = 1.2 * cm

    doc = SimpleDocTemplate(
        filepath,
        pagesize=PAGE,
        leftMargin=LEFT, rightMargin=RIGHT,
        topMargin=TOP, bottomMargin=BOTTOM,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title", parent=styles["Heading1"],
        fontSize=12, textColor=colors.HexColor("#1F59A3"),
        spaceAfter=3, spaceBefore=0,
    )
    sub_style = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=8, textColor=colors.grey, spaceAfter=6,
    )
    # Estilo para celdas con texto largo (hace wrap)
    cell_style = ParagraphStyle(
        "cell", parent=styles["Normal"],
        fontSize=6.5, leading=8, wordWrap='LTR',
    )
    cell_center = ParagraphStyle(
        "cell_c", parent=cell_style, alignment=1,
    )
    cell_right = ParagraphStyle(
        "cell_r", parent=cell_style, alignment=2,
    )
    header_style = ParagraphStyle(
        "hdr", parent=styles["Normal"],
        fontSize=6.5, leading=8, textColor=colors.white,
        fontName="Helvetica-Bold", alignment=1,
    )

    # Ancho total útil disponible (cm): 29.7 - 0.8*2 = 28.1 cm
    # Columnas: N°, Póliza, Recibo, Cliente, Cia, Ramo, Mon, Prima, Desde, Hasta, Ejec, Estado, RegPor, Hora
    # Totales:  0.8  2.5    2.5     5.5     3.0  3.5   1.6  2.6    2.2    2.2    3.0   2.2     2.5    1.5  = 36.1? reducir
    # Ajuste fino para que sumen ≤ 28.1 cm:
    col_w = [
        0.7*cm,   # N°
        2.4*cm,   # Póliza
        2.4*cm,   # Recibo
        5.0*cm,   # Cliente
        2.8*cm,   # Compañía
        3.2*cm,   # Ramo
        1.5*cm,   # Moneda
        2.4*cm,   # Prima Total
        2.0*cm,   # Vig Desde
        2.0*cm,   # Vig Hasta
        2.7*cm,   # Ejecutivo
        2.0*cm,   # Estado
        2.4*cm,   # Reg. por
        1.4*cm,   # Hora
    ]
    # suma = 34.5 → caben en 28.1 → escalar proporcionalmente
    available = PAGE[0] - LEFT - RIGHT
    scale = available / sum(col_w)
    col_w = [w * scale for w in col_w]

    pdf_headers = ["N°", "Póliza", "Recibo", "Cliente", "Compañía", "Ramo", "Moneda",
                   "Prima Total", "Vig. Desde", "Vig. Hasta", "Ejecutivo",
                   "Estado", "Reg. por", "Fecha/Hora"]
    pdf_col_idx = [0, 1, 2, 3, 4, 5, 7, 8, 9, 10, 11, 13, 14, 15]

    # Cabecera con Paragraph para wrap
    data = [[Paragraph(h, header_style) for h in pdf_headers]]

    for row in table_rows:
        def cell(i):
            if i == 8:
                return Paragraph(f"{row[i]:,.2f}", cell_right)
            elif i == 0:
                return Paragraph(str(row[i]), cell_center)
            else:
                return Paragraph(str(row[i]) if row[i] else "", cell_style)
        data.append([cell(i) for i in pdf_col_idx])

    table = Table(data, colWidths=col_w, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#1F59A3")),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
        ("TOPPADDING",    (0, 0), (-1, 0), 4),
        # Data
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF4FF")]),
        ("BOTTOMPADDING",  (0, 1), (-1, -1), 2),
        ("TOPPADDING",     (0, 1), (-1, -1), 2),
        # Grid
        ("GRID",           (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
    ]))

    total_prima = sum(row[8] for row in table_rows)
    total_label = Paragraph(
        f"<b>Total Prima: {total_prima:,.2f}</b> &nbsp;&nbsp; "
        f"<b>Total pólizas: {len(table_rows)}</b>",
        ParagraphStyle("tot", parent=styles["Normal"], fontSize=8,
                       textColor=colors.HexColor("#1F59A3"), spaceBefore=6),
    )

    story = [
        Paragraph("REPORTE DIARIO DE PÓLIZAS", title_style),
        Paragraph(f"Fecha: {today_str} — Generado: {datetime.utcnow().strftime('%H:%M:%S')} UTC", sub_style),
        table,
        total_label,
    ]
    doc.build(story)
    return filepath, filename






