from flask import request, session
from models.db import get_connection, get_encrypt_key
from datetime import datetime
from utils.rbac import Roles

def _get_poliza_activa_sql(alias='p'):
    return (
        f"COALESCE(NULLIF(TRIM(REPLACE(CONVERT({alias}.activo USING latin1), _latin1 0xA0, ' ')), ''), '0') = '1' "
        f"AND COALESCE(NULLIF(TRIM(REPLACE(CONVERT({alias}.anulado USING latin1), _latin1 0xA0, ' ')), ''), '0') = '0' "
        f"AND COALESCE({alias}.prima_anulada, 0) = 0"
    )

def _get_cuota_join_sql(poliza_alias='p', cuota_alias='q'):
    return f"""
        LEFT JOIN cuotas {cuota_alias}
          ON {cuota_alias}.idCuota = (
                SELECT q2.idCuota
                FROM cuotas q2
                WHERE q2.poliza_id = {poliza_alias}.idPoliza
                  AND q2.activo = 1
                ORDER BY
                    CASE
                        WHEN TRIM(COALESCE(
                                CONVERT(AES_DECRYPT(FROM_BASE64(q2.cupon), %s) USING utf8mb4),
                                CONVERT(AES_DECRYPT(q2.cupon, %s) USING utf8mb4),
                                CONVERT(q2.cupon USING utf8mb4)
                             )) COLLATE utf8mb4_0900_ai_ci =
                             TRIM(COALESCE(
                                CONVERT(AES_DECRYPT(FROM_BASE64({poliza_alias}.recibo), %s) USING utf8mb4),
                                CONVERT(AES_DECRYPT({poliza_alias}.recibo, %s) USING utf8mb4),
                                CONVERT({poliza_alias}.recibo USING utf8mb4)
                             )) COLLATE utf8mb4_0900_ai_ci
                        THEN 0
                        ELSE 1
                    END,
                    q2.fecha_vencimiento DESC,
                    q2.idCuota DESC
                LIMIT 1
          )
    """

def _dedupe_clientes_por_documento(clientes):
    """Conserva un solo cliente por tipo/nro de documento, priorizando el id mas reciente."""
    unicos = {}
    for cliente in clientes or []:
        tipo_doc = (cliente.get('tipo_documento') or '').strip().upper()
        numero_doc = (cliente.get('numero_documento') or '').strip().upper()
        if not numero_doc:
            key = f"ID:{cliente.get('idCliente')}"
        else:
            key = f"{tipo_doc}|{numero_doc}"

        actual = unicos.get(key)
        actual_id = int(actual.get('idCliente') or 0) if actual else 0
        nuevo_id = int(cliente.get('idCliente') or 0)
        if actual is None or nuevo_id > actual_id:
            unicos[key] = cliente

    return sorted(unicos.values(), key=lambda c: int(c.get('idCliente') or 0), reverse=True)

def _get_cliente_ids_relacionados(cur, key, cliente, es_subagente=False, usuario_actual=None):
    """Obtiene todos los idsCliente que comparten el mismo documento."""
    if not cliente:
        return []

    numero_documento = (cliente.get('numero_documento') or '').strip()
    tipo_documento = (cliente.get('tipo_documento') or '').strip()
    id_cliente = cliente.get('idCliente')

    if not numero_documento:
        return [id_cliente] if id_cliente else []

    query = """
        SELECT idCliente
        FROM clientes
        WHERE (
            CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR) = %s
            OR CAST(AES_DECRYPT(numero_documento, %s) AS CHAR) = %s
            OR numero_documento = %s
        )
    """
    params = [key, numero_documento, key, numero_documento, numero_documento]

    if tipo_documento:
        query += " AND tipo_documento = %s"
        params.append(tipo_documento)

    if es_subagente:
        query += " AND subagente = %s"
        params.append(usuario_actual)

    query += " ORDER BY idCliente DESC"
    cur.execute(query, params)
    rows = cur.fetchall() or []

    ids = []
    for row in rows:
        try:
            rid = int(row.get('idCliente'))
        except Exception:
            rid = None
        if rid and rid not in ids:
            ids.append(rid)

    if not ids and id_cliente:
        ids = [id_cliente]

    return ids

def get_estado_cuenta_data(filtros_input=None):
    """
    Obtiene los datos para el estado de cuenta de un cliente con filtros aplicados.
    Parámetros:
        filtros_input: dict con filtros (si viene de POST) o None para usar request.args (GET)
    Retorna: dict con 'cliente', 'polizas', 'totales', 'filtros_options'
    """
    try:
        # Obtener parámetros de filtro desde el parámetro o desde request.args
        if filtros_input:
            # Filtros vienen desde POST (pasados como parámetro)
            filters = {
                'cliente_id': filtros_input.get('cliente_id', ''),
                'cliente_search': filtros_input.get('cliente_search', ''),
                'tipo_documento': filtros_input.get('tipo_documento', ''),
                'numero_documento': filtros_input.get('numero_documento', ''),
                'compania': filtros_input.get('compania', ''),
                'moneda': filtros_input.get('moneda', ''),
                'ramo': filtros_input.get('ramo', ''),
                'estado': filtros_input.get('estado', ''),
                'fecha_desde': filtros_input.get('fecha_desde', ''),
                'fecha_hasta': filtros_input.get('fecha_hasta', '')
            }
        else:
            # Filtros vienen desde GET (request.args) - para compatibilidad
            filters = {
                'cliente_id': request.args.get('cliente_id', ''),
                'cliente_search': request.args.get('cliente_search', ''),
                'tipo_documento': request.args.get('tipo_documento', ''),
                'numero_documento': request.args.get('numero_documento', ''),
                'compania': request.args.get('compania', ''),
                'moneda': request.args.get('moneda', ''),
                'ramo': request.args.get('ramo', ''),
                'estado': request.args.get('estado', ''),
                'fecha_desde': request.args.get('fecha_desde', ''),
                'fecha_hasta': request.args.get('fecha_hasta', '')
            }

        if (filters.get('estado') or '').strip().upper() == 'CANCELADO':
            filters['estado'] = 'PAGADO'



        cnx = get_connection()
        cur = cnx.cursor(dictionary=True)
        key = get_encrypt_key()

        # Datos del cliente
        cliente = None
        polizas = []
        
        # RBAC: Verificar rol y usuario para filtros
        role_name = session.get('role_name')
        usuario_actual = session.get('user')
        es_subagente = (role_name == Roles.SUB_AGENTE)

        # Buscar cliente
        if filters['cliente_id']:
            query = """
                SELECT idCliente, 
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(razon_social), %s) AS CHAR), CAST(AES_DECRYPT(razon_social, %s) AS CHAR), razon_social) AS razon_social,
                       tipo_documento, 
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR), CAST(AES_DECRYPT(numero_documento, %s) AS CHAR), numero_documento) AS numero_documento,
                       direccion, 
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(telefono), %s) AS CHAR), CAST(AES_DECRYPT(telefono, %s) AS CHAR), telefono) AS telefono,
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(email), %s) AS CHAR), CAST(AES_DECRYPT(email, %s) AS CHAR), email) AS email,
                       subagente
                FROM clientes 
                WHERE idCliente = %s
            """
            params = [key, key, key, key, key, key, key, key, filters['cliente_id']]
            
            if es_subagente:
                query += " AND subagente = %s"
                params.append(usuario_actual)
                
            cur.execute(query, params)
            cliente = cur.fetchone()


        elif filters['tipo_documento'] and filters['numero_documento']:
            # Búsqueda por tipo y número de documento (sin necesidad de cliente_search)
            query = """
                SELECT idCliente, 
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(razon_social), %s) AS CHAR), CAST(AES_DECRYPT(razon_social, %s) AS CHAR), razon_social) AS razon_social,
                       tipo_documento, 
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR), CAST(AES_DECRYPT(numero_documento, %s) AS CHAR), numero_documento) AS numero_documento,
                       direccion, 
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(telefono), %s) AS CHAR), CAST(AES_DECRYPT(telefono, %s) AS CHAR), telefono) AS telefono,
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(email), %s) AS CHAR), CAST(AES_DECRYPT(email, %s) AS CHAR), email) AS email,
                       subagente
                FROM clientes 
                WHERE tipo_documento = %s 
                  AND (
                    CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR) = %s
                    OR CAST(AES_DECRYPT(numero_documento, %s) AS CHAR) = %s
                    OR numero_documento = %s
                  )
            """
            params = [key, key, key, key, key, key, key, key, filters['tipo_documento'], key, filters['numero_documento'], key, filters['numero_documento'], filters['numero_documento']]
            
            if es_subagente:
                query += " AND subagente = %s"
                params.append(usuario_actual)
                
            cur.execute(query, params)
            cliente = cur.fetchone()


        elif filters['numero_documento']:
            # Búsqueda solo por número de documento
            query = """
                SELECT idCliente, 
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(razon_social), %s) AS CHAR), CAST(AES_DECRYPT(razon_social, %s) AS CHAR), razon_social) AS razon_social,
                       tipo_documento, 
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR), CAST(AES_DECRYPT(numero_documento, %s) AS CHAR), numero_documento) AS numero_documento,
                       direccion, 
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(telefono), %s) AS CHAR), CAST(AES_DECRYPT(telefono, %s) AS CHAR), telefono) AS telefono,
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(email), %s) AS CHAR), CAST(AES_DECRYPT(email, %s) AS CHAR), email) AS email
                FROM clientes
                WHERE (
                    CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR) = %s
                    OR CAST(AES_DECRYPT(numero_documento, %s) AS CHAR) = %s
                    OR numero_documento = %s
                )
            """
            params = [key, key, key, key, key, key, key, key, key, filters['numero_documento'], key, filters['numero_documento'], filters['numero_documento']]
            
            if es_subagente:
                query += " AND subagente = %s"
                params.append(usuario_actual)
                
            cur.execute(query, params)
            cliente = cur.fetchone()


        elif filters['cliente_search']:
            # Búsqueda por texto en nombre o documento
            search_term = f"%{filters['cliente_search']}%"
            query = """
                SELECT idCliente, 
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(razon_social), %s) AS CHAR), CAST(AES_DECRYPT(razon_social, %s) AS CHAR), razon_social) AS razon_social,
                       tipo_documento, 
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR), CAST(AES_DECRYPT(numero_documento, %s) AS CHAR), numero_documento) AS numero_documento,
                       direccion, 
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(telefono), %s) AS CHAR), CAST(AES_DECRYPT(telefono, %s) AS CHAR), telefono) AS telefono,
                       COALESCE(CAST(AES_DECRYPT(FROM_BASE64(email), %s) AS CHAR), CAST(AES_DECRYPT(email, %s) AS CHAR), email) AS email
                FROM clientes 
                WHERE (
                    CAST(AES_DECRYPT(FROM_BASE64(razon_social), %s) AS CHAR) LIKE %s
                    OR CAST(AES_DECRYPT(razon_social, %s) AS CHAR) LIKE %s
                    OR razon_social LIKE %s
                    OR CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR) LIKE %s
                    OR CAST(AES_DECRYPT(numero_documento, %s) AS CHAR) LIKE %s
                    OR numero_documento LIKE %s
                )
            """
            params = [key, key, key, key, key, key, key, key, key, search_term, key, search_term, search_term, key, search_term, key, search_term, search_term]
            
            if es_subagente:
                query += " AND subagente = %s"
                params.append(usuario_actual)
                
            query += " LIMIT 1"
            
            cur.execute(query, params)
            cliente = cur.fetchone()



        if cliente and role_name == Roles.SUB_AGENTE:
            if cliente.get('subagente') and cliente.get('subagente') != usuario_actual:
                cliente = None

        if cliente:
            cliente_ids = _get_cliente_ids_relacionados(cur, key, cliente, es_subagente, usuario_actual)
            cliente_ids = cliente_ids or [cliente['idCliente']]
            placeholders = ', '.join(['%s'] * len(cliente_ids))

            query = f"""
                SELECT 
                    p.idPoliza,
                    p.cia AS compania,
                    p.ramo,
                    p.ramos_producto AS producto,
                    p.tipo_doc,
                    COALESCE(CAST(AES_DECRYPT(FROM_BASE64(p.poliza), %s) AS CHAR), CAST(AES_DECRYPT(p.poliza, %s) AS CHAR), p.poliza) AS poliza,
                    COALESCE(CAST(AES_DECRYPT(FROM_BASE64(p.recibo), %s) AS CHAR), CAST(AES_DECRYPT(p.recibo, %s) AS CHAR), p.recibo) AS proforma,
                    COALESCE(CAST(AES_DECRYPT(FROM_BASE64(q.cupon), %s) AS CHAR), CAST(AES_DECRYPT(q.cupon, %s) AS CHAR), q.cupon) AS cupon,
                    COALESCE(CAST(AES_DECRYPT(FROM_BASE64(q.factura), %s) AS CHAR), CAST(AES_DECRYPT(q.factura, %s) AS CHAR), q.factura) AS factura,
                    DATE_FORMAT(q.fecha_pago, '%%d/%%m/%%Y') AS fecha_pago,
                    DATE_FORMAT(p.fecha_emision, '%%d/%%m/%%Y') AS fecha_emision,
                    DATE_FORMAT(p.vig_desde, '%%d/%%m/%%Y') AS vig_inicio,
                    DATE_FORMAT(p.vig_hasta, '%%d/%%m/%%Y') AS vig_fin,
                    DATE_FORMAT(COALESCE(q.fecha_vencimiento, p.fecha_vencimiento), '%%d/%%m/%%Y') AS fecha_venc,
                    COALESCE(q.moneda, p.moneda) AS moneda,
                    p.prima_comercial_igv AS monto_cta_cobrar,
                    CASE 
                        WHEN q.idCuota IS NOT NULL THEN CASE WHEN q.fecha_pago IS NOT NULL THEN 0 ELSE q.importe END
                        ELSE CASE WHEN UPPER(IFNULL(p.estado,'')) IN ('CANCELADO', 'PAGADO') THEN 0 ELSE p.prima_comercial_igv END
                    END AS monto_cta_pagar,
                    CASE 
                        WHEN q.idCuota IS NOT NULL THEN CASE WHEN q.fecha_pago IS NOT NULL THEN 'PAGADO' ELSE 'PENDIENTE' END
                        ELSE p.estado
                    END AS estado
                FROM polizas p
                {_get_cuota_join_sql('p', 'q')}
                WHERE p.cliente_id IN ({placeholders})
                  AND {_get_poliza_activa_sql('p')}
            """

            params = [key, key, key, key, key, key, key, key, key, key, key, key, *cliente_ids]

            # Aplicar filtros adicionales
            if filters['compania']:
                query += " AND p.cia = %%s"
                params.append(filters['compania'])

            if filters['moneda']:
                moneda_filtro = filters['moneda'].upper()
                if 'SOLES' in moneda_filtro or 'S/' in moneda_filtro:
                    query += " AND (UPPER(p.moneda) LIKE '%%SOLES%%' OR UPPER(p.moneda) LIKE '%%S/%%')"
                elif 'DOLAR' in moneda_filtro or 'US$' in moneda_filtro or 'USD' in moneda_filtro:
                    query += " AND (UPPER(p.moneda) LIKE '%%DOLAR%%' OR UPPER(p.moneda) LIKE '%%US$%%' OR UPPER(p.moneda) LIKE '%%USD%%')"
                else:
                    query += " AND p.moneda = %%s"
                    params.append(filters['moneda'])

            if filters['ramo']:
                query += " AND p.ramo = %%s"
                params.append(filters['ramo'])

            # estado filtrado luego en Python para considerar cuotas

            if filters['fecha_desde'] and filters['fecha_hasta']:
                # Buscar pólizas donde CUALQUIER fecha esté dentro del rango
                query += """ AND (
                    (p.fecha_emision BETWEEN %%s AND %%s) OR
                    (p.fecha_vencimiento BETWEEN %%s AND %%s) OR
                    (p.vig_desde BETWEEN %%s AND %%s) OR
                    (p.vig_hasta BETWEEN %%s AND %%s) OR
                    (q.fecha_vencimiento BETWEEN %%s AND %%s)
                )"""
                params.extend([
                    filters['fecha_desde'], filters['fecha_hasta'],  # fecha_emision
                    filters['fecha_desde'], filters['fecha_hasta'],  # fecha_vencimiento
                    filters['fecha_desde'], filters['fecha_hasta'],  # vig_desde
                    filters['fecha_desde'], filters['fecha_hasta'],  # vig_hasta
                    filters['fecha_desde'], filters['fecha_hasta']   # cuotas.fecha_vencimiento
                ])
            elif filters['fecha_desde']:
                # Solo fecha desde: cualquier fecha >= fecha_desde
                query += """ AND (
                    p.fecha_emision >= %%s OR
                    p.fecha_vencimiento >= %%s OR
                    p.vig_desde >= %%s OR
                    p.vig_hasta >= %%s OR
                    q.fecha_vencimiento >= %%s
                )"""
                params.extend([
                    filters['fecha_desde'],
                    filters['fecha_desde'],
                    filters['fecha_desde'],
                    filters['fecha_desde'],
                    filters['fecha_desde']
                ])
            elif filters['fecha_hasta']:
                # Solo fecha hasta: cualquier fecha <= fecha_hasta
                query += """ AND (
                    p.fecha_emision <= %%s OR
                    p.fecha_vencimiento <= %%s OR
                    p.vig_desde <= %%s OR
                    p.vig_hasta <= %%s OR
                    q.fecha_vencimiento <= %%s
                )"""
                params.extend([
                    filters['fecha_hasta'],
                    filters['fecha_hasta'],
                    filters['fecha_hasta'],
                    filters['fecha_hasta'],
                    filters['fecha_hasta']
                ])

            query += " ORDER BY p.cia ASC, p.fecha_emision DESC, p.vig_desde DESC"

            # Reemplazar %% por % para la ejecución
            query_exec = query.replace('%%', '%')
            cur.execute(query_exec, params)
            polizas = cur.fetchall() or []

            if filters['estado']:
                est = filters['estado'].strip().upper()
                if est == 'PAGADO':
                    est = 'CANCELADO'
                polizas = [r for r in polizas if (r.get('estado') or '').upper() == est]

            # (No filtrar aquí: mostrar todas las pólizas, no eliminar por monto_cta_pagar)

        # Calcular totales por moneda
        totales = {
            'soles_cobrar': 0,
            'soles_pagar': 0,
            'dolares_cobrar': 0,
            'dolares_pagar': 0
        }

        for p in polizas:
            moneda_upper = p['moneda'].upper() if p['moneda'] else ''
            # Reconocer SOLES o S/
            if 'SOLES' in moneda_upper or 'S/' in moneda_upper:
                totales['soles_cobrar'] += float(p['monto_cta_cobrar'] or 0)
                totales['soles_pagar'] += float(p['monto_cta_pagar'] or 0)
            # Reconocer DÓLARES, DOLARES, US$ o USD
            elif 'DOLAR' in moneda_upper or 'US$' in moneda_upper or 'USD' in moneda_upper:
                totales['dolares_cobrar'] += float(p['monto_cta_cobrar'] or 0)
                totales['dolares_pagar'] += float(p['monto_cta_pagar'] or 0)

        # Obtener opciones para los filtros (sin depender del cliente)
        cur.execute("SELECT nombre_corto FROM companias WHERE nombre_corto IS NOT NULL AND nombre_corto != '' ORDER BY nombre_corto")
        companias = [row['nombre_corto'] for row in cur.fetchall()]

        cur.execute("SELECT DISTINCT nombre FROM ramos WHERE estado = 'Activo' ORDER BY nombre")
        ramos = [row['nombre'] for row in cur.fetchall()]

        # Obtener los estados reales de la tabla polizas
        cur.execute(
            f"""
            SELECT DISTINCT estado
            FROM polizas
            WHERE {_get_poliza_activa_sql('polizas')}
              AND estado IS NOT NULL
              AND estado != ''
              AND UPPER(estado) NOT IN ('INACTIVO', 'INACTIVA')
            ORDER BY estado
            """
        )
        estados_raw = [row['estado'] for row in cur.fetchall()]
        estados = []
        seen = set()
        for e in estados_raw:
            disp = 'PAGADO' if (e or '').strip().upper() == 'CANCELADO' else e
            if not disp:
                continue
            key_disp = disp.strip().upper()
            if key_disp in {'0', '1'}:
                continue
            if key_disp in seen:
                continue
            seen.add(key_disp)
            estados.append(disp)
        for requerido in ('PENDIENTE', 'PAGADO'):
            if requerido not in seen:
                estados.insert(0, requerido)
                seen.add(requerido)

        # Normalizar monedas para el filtro (mostrar opciones simples)
        monedas_normalizadas = ['SOLES', 'DÓLARES']

        filtros_options = {
            'companias': companias,
            'ramos': ramos,
            'estados': estados,
            'monedas': monedas_normalizadas
        }

        cur.close()
        cnx.close()

        return {
            'cliente': cliente,
            'polizas': polizas,
            'totales': totales,
            'filtros_options': filtros_options,
            'filtros_aplicados': filters
        }

    except Exception as e:
        print(f"Error en get_estado_cuenta_data: {e}")
        import traceback
        traceback.print_exc()
        return {
            'cliente': None,
            'polizas': [],
            'totales': {},
            'filtros_options': {'companias': [], 'ramos': [], 'estados': [], 'monedas': []},
            'filtros_aplicados': {}
        }


def buscar_clientes(search_term):
    """
    Busca clientes por nombre, RUC o DNI.
    Retorna lista de clientes encontrados.
    """
    try:
        cnx = get_connection()
        cur = cnx.cursor(dictionary=True)
        key = get_encrypt_key()

        search = f"%{search_term}%"
        
        # RBAC: Si es SUB AGENTE, filtrar por su usuario
        role_name = session.get('role_name')
        usuario_actual = session.get('user')
        
        if role_name == Roles.SUB_AGENTE:
            cur.execute("""
                SELECT 
                    idCliente,
                    COALESCE(CAST(AES_DECRYPT(FROM_BASE64(razon_social), %s) AS CHAR), CAST(AES_DECRYPT(razon_social, %s) AS CHAR), razon_social) AS razon_social,
                    tipo_documento,
                    COALESCE(CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR), CAST(AES_DECRYPT(numero_documento, %s) AS CHAR), numero_documento) AS numero_documento,
                    COALESCE(CAST(AES_DECRYPT(FROM_BASE64(telefono), %s) AS CHAR), CAST(AES_DECRYPT(telefono, %s) AS CHAR), telefono) AS telefono,
                    COALESCE(CAST(AES_DECRYPT(FROM_BASE64(email), %s) AS CHAR), CAST(AES_DECRYPT(email, %s) AS CHAR), email) AS email
                FROM clientes
                WHERE (
                    CAST(AES_DECRYPT(FROM_BASE64(razon_social), %s) AS CHAR) LIKE %s
                    OR CAST(AES_DECRYPT(razon_social, %s) AS CHAR) LIKE %s
                    OR razon_social LIKE %s
                    OR CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR) LIKE %s
                    OR CAST(AES_DECRYPT(numero_documento, %s) AS CHAR) LIKE %s
                    OR numero_documento LIKE %s
                )
                  AND subagente = %s
                ORDER BY idCliente DESC, razon_social
                LIMIT 20
            """, (key, key, key, key, key, key, key, key, key, search, key, search, search, key, search, key, search, search, usuario_actual))
        else:
            cur.execute("""
                SELECT 
                    idCliente,
                    COALESCE(CAST(AES_DECRYPT(FROM_BASE64(razon_social), %s) AS CHAR), CAST(AES_DECRYPT(razon_social, %s) AS CHAR), razon_social) AS razon_social,
                    tipo_documento,
                    COALESCE(CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR), CAST(AES_DECRYPT(numero_documento, %s) AS CHAR), numero_documento) AS numero_documento,
                    COALESCE(CAST(AES_DECRYPT(FROM_BASE64(telefono), %s) AS CHAR), CAST(AES_DECRYPT(telefono, %s) AS CHAR), telefono) AS telefono,
                    COALESCE(CAST(AES_DECRYPT(FROM_BASE64(email), %s) AS CHAR), CAST(AES_DECRYPT(email, %s) AS CHAR), email) AS email
                FROM clientes
                WHERE (
                    CAST(AES_DECRYPT(FROM_BASE64(razon_social), %s) AS CHAR) LIKE %s
                    OR CAST(AES_DECRYPT(razon_social, %s) AS CHAR) LIKE %s
                    OR razon_social LIKE %s
                    OR CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR) LIKE %s
                    OR CAST(AES_DECRYPT(numero_documento, %s) AS CHAR) LIKE %s
                    OR numero_documento LIKE %s
                )
                ORDER BY idCliente DESC, razon_social
                LIMIT 20
            """, (key, key, key, key, key, key, key, key, key, search, key, search, search, key, search, key, search, search))

        clientes = _dedupe_clientes_por_documento(cur.fetchall() or [])

        cur.close()
        cnx.close()

        return clientes

    except Exception as e:
        print(f"Error en buscar_clientes: {e}")
        return []


def export_estado_cuenta_data(args, fmt='xlsx'):
    """Genera un archivo (xlsx o pdf) con las pólizas filtradas.
    args: request.args-like (mapping)
    fmt: 'xlsx' o 'pdf'
    Retorna: (filepath, filename)
    """
    from models.db import get_connection
    import os
    from datetime import datetime

    # Reconstruir filtros desde args
    filters = {
        'cliente_id': args.get('cliente_id', ''),
        'cliente_search': args.get('cliente_search', ''),
        'tipo_documento': args.get('tipo_documento', ''),
        'numero_documento': args.get('numero_documento', ''),
        'compania': args.get('compania', ''),
        'moneda': args.get('moneda', ''),
        'ramo': args.get('ramo', ''),
        'estado': args.get('estado', ''),
        'fecha_desde': args.get('fecha_desde', ''),
        'fecha_hasta': args.get('fecha_hasta', '')
    }

    if (filters.get('estado') or '').strip().upper() == 'CANCELADO':
        filters['estado'] = 'PAGADO'

    # Reusar la lógica de consulta para obtener polizas
    cnx = get_connection()
    cur = cnx.cursor(dictionary=True)
    key = get_encrypt_key()

    cliente = None
    polizas = []

    # RBAC: Verificar rol y usuario para filtros
    role_name = session.get('role_name')
    usuario_actual = session.get('user')
    es_subagente = (role_name == Roles.SUB_AGENTE)

    # Buscar cliente (misma lógica que get_estado_cuenta_data)
    if filters['cliente_id']:
        query = """
            SELECT idCliente, 
                   COALESCE(CAST(AES_DECRYPT(FROM_BASE64(razon_social), %s) AS CHAR), CAST(AES_DECRYPT(razon_social, %s) AS CHAR), razon_social) AS razon_social,
                   tipo_documento, 
                   COALESCE(CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR), CAST(AES_DECRYPT(numero_documento, %s) AS CHAR), numero_documento) AS numero_documento,
                   subagente
            FROM clientes WHERE idCliente = %s
        """
        params = [key, key, key, key, filters['cliente_id']]
        
        if es_subagente:
            query += " AND subagente = %s"
            params.append(usuario_actual)
            
        cur.execute(query, params)
        cliente = cur.fetchone()

    elif filters['tipo_documento'] and filters['numero_documento']:
        query = """
            SELECT idCliente, 
                   COALESCE(CAST(AES_DECRYPT(FROM_BASE64(razon_social), %s) AS CHAR), CAST(AES_DECRYPT(razon_social, %s) AS CHAR), razon_social) AS razon_social,
                   tipo_documento, 
                   COALESCE(CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR), CAST(AES_DECRYPT(numero_documento, %s) AS CHAR), numero_documento) AS numero_documento,
                   subagente
            FROM clientes 
            WHERE tipo_documento = %s 
              AND (
                CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR) = %s
                OR CAST(AES_DECRYPT(numero_documento, %s) AS CHAR) = %s
                OR numero_documento = %s
              )
        """
        params = [key, key, key, key, filters['tipo_documento'], key, filters['numero_documento'], key, filters['numero_documento'], filters['numero_documento']]
        
        if es_subagente:
            query += " AND subagente = %s"
            params.append(usuario_actual)
            
        cur.execute(query, params)
        cliente = cur.fetchone()

    elif filters['numero_documento']:
        query = """
            SELECT idCliente, 
                   COALESCE(CAST(AES_DECRYPT(FROM_BASE64(razon_social), %s) AS CHAR), CAST(AES_DECRYPT(razon_social, %s) AS CHAR), razon_social) AS razon_social,
                   tipo_documento, 
                   COALESCE(CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR), CAST(AES_DECRYPT(numero_documento, %s) AS CHAR), numero_documento) AS numero_documento,
                   subagente
            FROM clientes 
            WHERE (
                CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR) = %s
                OR CAST(AES_DECRYPT(numero_documento, %s) AS CHAR) = %s
                OR numero_documento = %s
            )
        """
        params = [key, key, key, key, key, filters['numero_documento'], key, filters['numero_documento'], filters['numero_documento']]
        
        if es_subagente:
            query += " AND subagente = %s"
            params.append(usuario_actual)
            
        cur.execute(query, params)
        cliente = cur.fetchone()

    elif filters['cliente_search']:
        search_term = f"%{filters['cliente_search']}%"
        query = """
            SELECT idCliente, 
                   COALESCE(CAST(AES_DECRYPT(FROM_BASE64(razon_social), %s) AS CHAR), CAST(AES_DECRYPT(razon_social, %s) AS CHAR), razon_social) AS razon_social,
                   tipo_documento, 
                   COALESCE(CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR), CAST(AES_DECRYPT(numero_documento, %s) AS CHAR), numero_documento) AS numero_documento,
                   subagente
            FROM clientes 
            WHERE (
                CAST(AES_DECRYPT(FROM_BASE64(razon_social), %s) AS CHAR) LIKE %s
                OR CAST(AES_DECRYPT(razon_social, %s) AS CHAR) LIKE %s
                OR razon_social LIKE %s
                OR CAST(AES_DECRYPT(FROM_BASE64(numero_documento), %s) AS CHAR) LIKE %s
                OR CAST(AES_DECRYPT(numero_documento, %s) AS CHAR) LIKE %s
                OR numero_documento LIKE %s
            )
        """
        params = [key, key, key, key, key, search_term, key, search_term, search_term, key, search_term, key, search_term, search_term]
        
        if es_subagente:
            query += " AND subagente = %s"
            params.append(usuario_actual)
            
        query += " LIMIT 1"
        cur.execute(query, params)
        cliente = cur.fetchone()

    if cliente and role_name == Roles.SUB_AGENTE:
        if cliente.get('subagente') and cliente.get('subagente') != usuario_actual:
            cliente = None

    if cliente:
        cliente_ids = _get_cliente_ids_relacionados(cur, key, cliente, es_subagente, usuario_actual)
        cliente_ids = cliente_ids or [cliente['idCliente']]
        placeholders = ', '.join(['%s'] * len(cliente_ids))
        query = f"""
            SELECT 
                p.cia AS compania,
                p.ramo,
                p.ramos_producto AS producto,
                p.tipo_doc,
                COALESCE(CAST(AES_DECRYPT(FROM_BASE64(p.poliza), %s) AS CHAR), CAST(AES_DECRYPT(p.poliza, %s) AS CHAR), p.poliza) AS poliza,
                COALESCE(CAST(AES_DECRYPT(FROM_BASE64(p.recibo), %s) AS CHAR), CAST(AES_DECRYPT(p.recibo, %s) AS CHAR), p.recibo) AS proforma,
                COALESCE(CAST(AES_DECRYPT(FROM_BASE64(q.cupon), %s) AS CHAR), CAST(AES_DECRYPT(q.cupon, %s) AS CHAR), q.cupon) AS cupon,
                COALESCE(CAST(AES_DECRYPT(FROM_BASE64(q.factura), %s) AS CHAR), CAST(AES_DECRYPT(q.factura, %s) AS CHAR), q.factura) AS factura,
                DATE_FORMAT(q.fecha_pago, '%%d/%%m/%%Y') AS fecha_pago,
                DATE_FORMAT(p.fecha_emision, '%%d/%%m/%%Y') AS fecha_emision,
                DATE_FORMAT(p.vig_desde, '%%d/%%m/%%Y') AS vig_inicio,
                DATE_FORMAT(p.vig_hasta, '%%d/%%m/%%Y') AS vig_fin,
                DATE_FORMAT(COALESCE(q.fecha_vencimiento, p.fecha_vencimiento), '%%d/%%m/%%Y') AS fecha_venc,
                COALESCE(q.moneda, p.moneda) AS moneda,
                p.prima_comercial_igv AS monto_cta_cobrar,
                CASE 
                    WHEN q.idCuota IS NOT NULL THEN CASE WHEN q.fecha_pago IS NOT NULL THEN 0 ELSE q.importe END
                    ELSE CASE WHEN UPPER(IFNULL(p.estado,'')) IN ('CANCELADO', 'PAGADO') THEN 0 ELSE p.prima_comercial_igv END
                END AS monto_cta_pagar,
                CASE 
                    WHEN q.idCuota IS NOT NULL THEN CASE WHEN q.fecha_pago IS NOT NULL THEN 'PAGADO' ELSE 'PENDIENTE' END
                    ELSE p.estado
                END AS estado
            FROM polizas p
            {_get_cuota_join_sql('p', 'q')}
            WHERE p.cliente_id IN ({placeholders})
              AND {_get_poliza_activa_sql('p')}
        """
        params = [key, key, key, key, key, key, key, key, key, key, key, key, *cliente_ids]

        if filters['compania']:
            query += " AND p.cia = %%s"
            params.append(filters['compania'])

        if filters['moneda']:
            moneda_filtro = filters['moneda'].upper()
            if 'SOLES' in moneda_filtro or 'S/' in moneda_filtro:
                query += " AND (UPPER(p.moneda) LIKE '%%SOLES%%' OR UPPER(p.moneda) LIKE '%%S/%%')"
            elif 'DOLAR' in moneda_filtro or 'US$' in moneda_filtro or 'USD' in moneda_filtro:
                query += " AND (UPPER(p.moneda) LIKE '%%DOLAR%%' OR UPPER(p.moneda) LIKE '%%US$%%' OR UPPER(p.moneda) LIKE '%%USD%%')"
            else:
                query += " AND p.moneda = %%s"
                params.append(filters['moneda'])

        if filters['ramo']:
            query += " AND p.ramo = %%s"
            params.append(filters['ramo'])

        # estado filtrado luego en Python para considerar cuotas

        if filters['fecha_desde'] and filters['fecha_hasta']:
            query += """ AND ((p.fecha_emision BETWEEN %%s AND %%s) OR (p.fecha_vencimiento BETWEEN %%s AND %%s) OR (p.vig_desde BETWEEN %%s AND %%s) OR (p.vig_hasta BETWEEN %%s AND %%s) OR (q.fecha_vencimiento BETWEEN %%s AND %%s))"""
            params.extend([filters['fecha_desde'], filters['fecha_hasta']] * 5)
        elif filters['fecha_desde']:
            query += """ AND (p.fecha_emision >= %%s OR p.fecha_vencimiento >= %%s OR p.vig_desde >= %%s OR p.vig_hasta >= %%s OR q.fecha_vencimiento >= %%s)"""
            params.extend([filters['fecha_desde']] * 5)
        elif filters['fecha_hasta']:
            query += """ AND (p.fecha_emision <= %%s OR p.fecha_vencimiento <= %%s OR p.vig_desde <= %%s OR p.vig_hasta <= %%s OR q.fecha_vencimiento <= %%s)"""
            params.extend([filters['fecha_hasta']] * 5)

        query += " ORDER BY p.cia ASC, p.fecha_emision DESC, p.vig_desde DESC"
        
        # Reemplazar %% por % para la ejecución
        query_exec = query.replace('%%', '%')
        cur.execute(query_exec, params)
        polizas = cur.fetchall() or []

        if filters['estado']:
            est = filters['estado'].strip().upper()
            if est == 'PAGADO':
                est = 'CANCELADO'
            polizas = [r for r in polizas if (r.get('estado') or '').upper() == est]

            # (No filtrar aquí para export: mantener todas las pólizas, no eliminar por monto_cta_pagar)
        cur.close()
    cnx.close()

    # Preparar datos para exportar
    headers = [
        'Compañía',
        'Ramo',
        'Producto',
        'Tipo Doc',
        'N° de Póliza',
        'Proforma',
        'Cupón',
        'Fecha Emisión',
        'Vigencia (Desde - Hasta)',
        'Factura',
        'Fecha de Pago',
        'Fecha Vencimiento',
        'Moneda',
        'Cta. Cobrar',
        'Cta. Pagar',
        'Estado'
    ]
    rows = []
    for p in polizas:
        vig = f"{p.get('vig_inicio') or '-'} - {p.get('vig_fin') or '-'}"
        moneda = p.get('moneda') or ''
        if moneda:
            mu = moneda.upper()
            if 'SOLES' in mu or 'S/' in mu:
                moneda = 'S/'
            elif 'DOLAR' in mu or 'US$' in mu or 'USD' in mu:
                moneda = 'US$'

        rows.append([
            p.get('compania') or '-',
            p.get('ramo') or '-',
            p.get('producto') or '-',
            p.get('tipo_doc') or '-',
            p.get('poliza') or '-',
            p.get('proforma') or '-',
            p.get('cupon') or '-',
            p.get('fecha_emision') or '-',
            vig,
            p.get('factura') or '-',
            p.get('fecha_pago') or '-',
            p.get('fecha_venc') or '-',
            moneda or '-',
            float(p.get('monto_cta_cobrar') or 0),
            float(p.get('monto_cta_pagar') or 0),
            'PAGADO' if (p.get('estado') or '').strip().upper() == 'CANCELADO' else (p.get('estado') or '-')
        ])

    # Crear carpeta de export
    upload_folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), '..', 'uploads')
    # Prefer current_app if disponible
    try:
        from flask import current_app
        upload_folder = current_app.config.get('UPLOAD_FOLDER', upload_folder)
    except Exception:
        pass

    exports_dir = os.path.join(upload_folder, 'exports')
    os.makedirs(exports_dir, exist_ok=True)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    base_name = f"estado_cuenta_{cliente['idCliente'] if cliente else 'sincliente'}_{ts}"

    if fmt == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = 'Estado de Cuenta'

            # No insertar logo (solicitado). Mantener layout simple.
            start_row = 1

            # Título (centrado sobre las 16 columnas)
            title_row = start_row
            title_cell = ws.cell(row=title_row, column=1, value='Estado de Cuenta')
            title_cell.font = Font(size=14, bold=True, color='1F59A3')
            ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=16)
            title_cell.alignment = Alignment(horizontal='center')

            # Información del cliente debajo del título
            info_row = title_row + 1
            if cliente:
                razon = (cliente.get('razon_social') or 'N/A')
                tipo_doc = cliente.get('tipo_documento') or ''
                num_doc = cliente.get('numero_documento') or 'N/A'
                info_text = f"Cliente: {razon}    Documento: {tipo_doc} - {num_doc}"
                ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=16)
                c = ws.cell(row=info_row, column=1, value=info_text)
                c.font = Font(size=9)
                c.alignment = Alignment(horizontal='left')

            # Encabezados
            header_row = title_row + 3
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=col_idx, value=h)
                cell.font = Font(bold=True, size=9, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='4472C4')
                cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
                thin = Side(border_style='thin', color='AAAAAA')
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # Ajuste de anchos de columna (más espacio para Producto y Vigencia)
            # Aumentados para evitar que el texto se desplace o queden espacios en blanco
            col_widths_chars = [30, 14, 30, 12, 18, 14, 12, 12, 36, 14, 14, 14, 10, 18, 18, 14]
            for i, w in enumerate(col_widths_chars, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            # Escribir filas con formato y acumular totales por moneda
            row_idx = header_row + 1
            totales_pdf = {}  # { 'S/': [cobrar, pagar], 'US$': [cobrar,pagar] }

            for r in rows:
                for col_idx, val in enumerate(r, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # Sanitizar cadenas: colapsar espacios múltiples y eliminar saltos de línea innecesarios
                    if isinstance(val, (float, int)):
                        cell.value = float(val)
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.font = Font(bold=True)
                    else:
                        # colapsar espacios (ej: 'A   B  C' -> 'A B C') y limpiar
                        txt = ' '.join(str(val).split()) if val is not None else '-'
                        cell.value = txt
                        # Fechas y moneda centradas
                        if col_idx in (8, 11, 12):  # fecha_emision (8), fecha_pago(11), fecha_venc(12)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif col_idx == 13:  # Moneda
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            # Wrap para evitar que el texto genere espacios vacíos; no usar shrink_to_fit
                            cell.alignment = Alignment(horizontal='left', wrap_text=True, vertical='center')

                # Fijar altura de fila para evitar desalineaciones (p.ej. 18pt)
                try:
                    ws.row_dimensions[row_idx].height = 18
                except Exception:
                    pass

                # Acumular totales (índices 12=moneda, 13=cobrar, 14=pagar en 0-based r[12], r[13], r[14])
                mon = ' '.join(str(r[12]).split()) if r[12] else '-'
                cobrar_val = float(r[13]) if isinstance(r[13], (float, int)) else 0.0
                pagar_val = float(r[14]) if isinstance(r[14], (float, int)) else 0.0
                if mon not in totales_pdf:
                    totales_pdf[mon] = [0.0, 0.0]
                totales_pdf[mon][0] += cobrar_val
                totales_pdf[mon][1] += pagar_val

                # Alternar color de fila (sutil)
                fill_color = 'FFFFFF' if (row_idx - header_row) % 2 == 1 else 'F7F9FC'
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill('solid', fgColor=fill_color)
                row_idx += 1

            # Ajustar altura del encabezado y título para balance visual
            try:
                ws.row_dimensions[header_row].height = 20
                ws.row_dimensions[title_row].height = 22
            except Exception:
                pass

            # Agregar filas de totales por moneda (estilo simple y legible)
            for mon_key, (cobrar, pagar) in sorted(totales_pdf.items()):
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=12)
                lbl_cell = ws.cell(row=row_idx, column=1, value=f'TOTAL {mon_key}')
                lbl_cell.font = Font(bold=True, color='1A3A6B')
                lbl_cell.alignment = Alignment(horizontal='right', vertical='center')

                mon_cell = ws.cell(row=row_idx, column=13, value=mon_key)
                mon_cell.font = Font(bold=True, color='1A3A6B')
                mon_cell.alignment = Alignment(horizontal='center', vertical='center')

                cobrar_cell = ws.cell(row=row_idx, column=14, value=cobrar)
                cobrar_cell.number_format = '#,##0.00'
                cobrar_cell.font = Font(bold=True, color='0047AB')
                cobrar_cell.alignment = Alignment(horizontal='right', vertical='center')

                pagar_cell = ws.cell(row=row_idx, column=15, value=pagar)
                pagar_cell.number_format = '#,##0.00'
                pagar_cell.font = Font(bold=True, color='B22222')
                pagar_cell.alignment = Alignment(horizontal='right', vertical='center')

                # Fondo sutil para fila de totales y borde superior
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill('solid', fgColor='FFF3CD')
                thick_side = Side(border_style='medium', color='4472C4')
                for col_idx in range(1, len(headers) + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.border = Border(top=thick_side)
                # altura de fila total
                try:
                    ws.row_dimensions[row_idx].height = 18
                except Exception:
                    pass
                row_idx += 1

            # Freeze header
            ws.freeze_panes = ws['A' + str(header_row + 1)]

            filename = f"{base_name}.xlsx"
            filepath = os.path.join(exports_dir, filename)
            wb.save(filepath)
            return filepath, filename
        except Exception as e:
            raise
    elif fmt == 'pdf':
        # Generar PDF con reportlab - formato optimizado para estado de cuenta
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

            filename = f"{base_name}.pdf"
            filepath = os.path.join(exports_dir, filename)

            # Landscape A4: ancho=297mm, alto=210mm. Márgenes 10mm c/u => usable=277mm
            PAGE_W, PAGE_H = landscape(A4)
            MARGIN = 10 * mm
            USABLE_W = PAGE_W - 2 * MARGIN

            doc = SimpleDocTemplate(
                filepath,
                pagesize=landscape(A4),
                rightMargin=MARGIN,
                leftMargin=MARGIN,
                topMargin=MARGIN,
                bottomMargin=MARGIN
            )
            elements = []
            styles = getSampleStyleSheet()

            # ── Logo con proporción real ──────────────────────────────────────
            try:
                from PIL import Image as PILImage
                logo_path = os.path.join(upload_folder, 'logo', 'Logo-banner.png')
                if os.path.exists(logo_path):
                    with PILImage.open(logo_path) as pil_img:
                        orig_w, orig_h = pil_img.size
                    MAX_W = 55 * mm
                    MAX_H = 18 * mm
                    ratio = min(MAX_W / orig_w, MAX_H / orig_h)
                    logo = Image(logo_path, width=orig_w * ratio, height=orig_h * ratio)
                    logo.hAlign = 'LEFT'
                    elements.append(logo)
                    elements.append(Spacer(1, 4 * mm))
            except Exception as e:
                print(f"[WARN] No se pudo cargar el logo: {e}")

            # ── Título principal ──────────────────────────────────────────────
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=14,
                textColor=colors.HexColor('#1F59A3'),
                spaceAfter=6,
                alignment=TA_CENTER
            )
            elements.append(Paragraph('<b>Estado de Cuenta</b>', title_style))

            # ── Info del cliente ──────────────────────────────────────────────
            if cliente:
                cliente_info_style = ParagraphStyle(
                    'ClienteInfo',
                    parent=styles['Normal'],
                    fontSize=9,
                    spaceAfter=4,
                    alignment=TA_LEFT
                )
                razon = (cliente.get('razon_social') or 'N/A').upper()
                tipo_doc = cliente.get('tipo_documento') or ''
                num_doc = cliente.get('numero_documento') or 'N/A'
                info_text = (
                    f"<b>Cliente:</b> {razon} &nbsp;&nbsp;|&nbsp;&nbsp; "
                    f"<b>Documento:</b> {tipo_doc} - {num_doc}"
                )
                elements.append(Paragraph(info_text, cliente_info_style))

            elements.append(Spacer(1, 4 * mm))

            # ── Estilos de párrafo para celdas (wrapping) ────────────────────
            cell_style = ParagraphStyle(
                'CellNormal',
                parent=styles['Normal'],
                fontSize=6.5,
                leading=8,
                wordWrap='CJK',
            )
            cell_bold = ParagraphStyle(
                'CellBold',
                parent=cell_style,
                fontName='Helvetica-Bold',
            )
            cell_blue = ParagraphStyle(
                'CellBlue',
                parent=cell_style,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#0066CC'),
                alignment=TA_RIGHT,
            )
            cell_red = ParagraphStyle(
                'CellRed',
                parent=cell_style,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#CC0000'),
                alignment=TA_RIGHT,
            )
            cell_center = ParagraphStyle(
                'CellCenter',
                parent=cell_style,
                alignment=TA_CENTER,
            )
            hdr_style = ParagraphStyle(
                'HdrStyle',
                parent=styles['Normal'],
                fontSize=6.5,
                leading=8,
                fontName='Helvetica-Bold',
                textColor=colors.whitesmoke,
                alignment=TA_CENTER,
                wordWrap='CJK',
            )

            # ── Anchos de columna (suman exactamente USABLE_W) ───────────────
            # 16 columnas: Compañía Ramo Producto TipoDoc NoPol Proforma Cupón
            #              FechaEm Vigencia Factura FechaPago FechaVenc Moneda CtaCob CtaPag Estado
            col_widths = [
                USABLE_W * 0.085,   # Compañía
                USABLE_W * 0.048,   # Ramo
                USABLE_W * 0.075,   # Producto
                USABLE_W * 0.055,   # Tipo Doc
                USABLE_W * 0.075,   # N° de Póliza
                USABLE_W * 0.065,   # Proforma
                USABLE_W * 0.055,   # Cupón
                USABLE_W * 0.062,   # Fecha Emisión
                USABLE_W * 0.095,   # Vigencia (Desde - Hasta)
                USABLE_W * 0.065,   # Factura
                USABLE_W * 0.062,   # Fecha de Pago
                USABLE_W * 0.062,   # Fecha Vencimiento
                USABLE_W * 0.040,   # Moneda
                USABLE_W * 0.060,   # Cta. Cobrar
                USABLE_W * 0.060,   # Cta. Pagar
                USABLE_W * 0.060,   # Estado
            ]
            # Ajuste fino: asegurar que la suma sea exactamente USABLE_W
            diff = USABLE_W - sum(col_widths)
            col_widths[8] += diff   # absorber diferencia en columna Vigencia

            # ── Encabezados como Paragraph para wrapping ─────────────────────
            hdr_row = [Paragraph(h, hdr_style) for h in headers]

            # ── Filas de datos ────────────────────────────────────────────────
            table_data = [hdr_row]

            # Acumuladores de totales por moneda  { moneda_key: [cobrar, pagar] }
            totales_pdf = {}   # e.g. {'S/': [100.0, 80.0], 'US$': [50.0, 40.0]}

            for r in rows:
                row_cells = []
                for i, val in enumerate(r):
                    if isinstance(val, float):
                        txt = "{:.2f}".format(val)
                        st = cell_blue if i == 13 else cell_red  # 13=CtaCob, 14=CtaPag
                    elif i in (7, 10, 11):  # fechas → centrado
                        txt = str(val) if val else '-'
                        st = cell_center
                    elif i == 12:  # moneda → centrado
                        txt = str(val) if val else '-'
                        st = cell_center
                    elif i == 15:  # estado → centrado
                        txt = str(val) if val else '-'
                        st = cell_center
                    else:
                        txt = str(val) if val else '-'
                        st = cell_style
                    row_cells.append(Paragraph(txt, st))
                table_data.append(row_cells)

                # Acumular totales (índice 12=moneda, 13=cobrar, 14=pagar)
                mon_key = str(r[12]) if r[12] else '-'
                if mon_key not in totales_pdf:
                    totales_pdf[mon_key] = [0.0, 0.0]
                totales_pdf[mon_key][0] += float(r[13]) if isinstance(r[13], float) else 0.0
                totales_pdf[mon_key][1] += float(r[14]) if isinstance(r[14], float) else 0.0

            # ── Filas de totales por moneda ───────────────────────────────────
            total_row_style_lbl = ParagraphStyle(
                'TotalLbl',
                parent=styles['Normal'],
                fontSize=7.5,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#1A3A6B'),
                alignment=TA_RIGHT,
            )
            total_row_style_blue = ParagraphStyle(
                'TotalBlue',
                parent=styles['Normal'],
                fontSize=7.5,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#0047AB'),
                alignment=TA_RIGHT,
            )
            total_row_style_red = ParagraphStyle(
                'TotalRed',
                parent=styles['Normal'],
                fontSize=7.5,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#B22222'),
                alignment=TA_RIGHT,
            )
            total_row_style_mon = ParagraphStyle(
                'TotalMon',
                parent=styles['Normal'],
                fontSize=7.5,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#1A3A6B'),
                alignment=TA_CENTER,
            )

            NUM_COLS = len(headers)   # 16
            total_rows_start = len(table_data)  # primera fila de totales (0-based)

            for mon_key, (cobrar, pagar) in sorted(totales_pdf.items()):
                total_cells = [Paragraph('', cell_style)] * NUM_COLS
                # Etiqueta "TOTAL" fusionada visualmente en col 0..11
                lbl_cell = Paragraph(f'TOTAL {mon_key}', total_row_style_lbl)
                total_cells = (
                    [lbl_cell] +
                    [Paragraph('', cell_style)] * 11 +   # cols 1-11 vacías
                    [Paragraph(mon_key,            total_row_style_mon),   # col 12 moneda
                     Paragraph(f'{cobrar:,.2f}',   total_row_style_blue),  # col 13 cobrar
                     Paragraph(f'{pagar:,.2f}',    total_row_style_red),   # col 14 pagar
                     Paragraph('',                 cell_style)]             # col 15 estado
                )
                table_data.append(total_cells)

            total_rows_end = len(table_data) - 1  # última fila de totales (0-based)

            t = Table(table_data, colWidths=col_widths, repeatRows=1)

            table_style_cmds = [
                # ── Encabezado ──
                ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('VALIGN',        (0, 0), (-1, 0), 'MIDDLE'),
                ('TOPPADDING',    (0, 0), (-1, 0), 5),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                ('LEFTPADDING',   (0, 0), (-1, 0), 3),
                ('RIGHTPADDING',  (0, 0), (-1, 0), 3),
                # ── Datos ──
                ('VALIGN',        (0, 1), (-1, -1), 'MIDDLE'),
                ('TOPPADDING',    (0, 1), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
                ('LEFTPADDING',   (0, 1), (-1, -1), 3),
                ('RIGHTPADDING',  (0, 1), (-1, -1), 3),
                # ── Grid y colores alternos ──
                ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#AAAAAA')),
                ('ROWBACKGROUNDS', (0, 1), (NUM_COLS - 1, total_rows_start - 1),
                                  [colors.white, colors.HexColor('#EEF2FA')]),
            ]

            # Estilo especial para las filas de totales
            for tr in range(total_rows_start, total_rows_end + 1):
                table_style_cmds += [
                    ('BACKGROUND',    (0, tr), (-1, tr), colors.HexColor('#FFF3CD')),
                    ('TOPPADDING',    (0, tr), (-1, tr), 5),
                    ('BOTTOMPADDING', (0, tr), (-1, tr), 5),
                    # Borde superior más grueso para separar visualmente del resto
                    ('LINEABOVE',     (0, tr), (-1, tr), 1.5, colors.HexColor('#4472C4')),
                    # Fusionar etiqueta "TOTAL" sobre las primeras 12 cols
                    ('SPAN',          (0, tr), (11, tr)),
                    ('ALIGN',         (0, tr), (11, tr), 'RIGHT'),
                ]

            t.setStyle(TableStyle(table_style_cmds))

            elements.append(t)

            # ── Footer ───────────────────────────────────────────────────────
            elements.append(Spacer(1, 5 * mm))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=7,
                textColor=colors.grey,
                alignment=TA_RIGHT
            )
            fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            elements.append(Paragraph(f'Generado: {fecha_generacion}', footer_style))

            doc.build(elements)
            return filepath, filename
        except Exception as e:
            raise
    else:
        raise ValueError('Formato no soportado')
