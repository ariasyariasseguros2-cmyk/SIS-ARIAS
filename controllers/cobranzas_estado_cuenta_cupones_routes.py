from flask import Blueprint, request, session, jsonify, current_app, send_file
from models.db import get_connection
from utils.rbac import Roles
import os


bp = Blueprint("cobranzas_estado_cuenta_cupones", __name__)


def _get_multi(name: str):
    values = request.args.getlist(name) or []
    out = []
    seen = set()
    for v in values:
        if v is None:
            continue
        for part in str(v).split(","):
            p = part.strip()
            if not p:
                continue
            if p in seen:
                continue
            seen.add(p)
            out.append(p)
    return out


def _validate_required(fecha_desde: str, fecha_hasta: str):
    missing = []
    if not fecha_desde:
        missing.append("Del")
    if not fecha_hasta:
        missing.append("Al")
    return missing


def _fetch_rows():
    fecha_desde = (request.args.get("fecha_desde") or "").strip()
    fecha_hasta = (request.args.get("fecha_hasta") or "").strip()
    poliza_cupon = (request.args.get("poliza_cupon") or "").strip()
    cliente_ids = _get_multi("cliente_id")
    cias = _get_multi("cia")
    ramos = _get_multi("ramo")
    ejecutivos = _get_multi("ejecutivo")
    sub_agentes = _get_multi("sub_agente")
    estados = [e.upper() for e in _get_multi("estado")]

    missing = _validate_required(fecha_desde, fecha_hasta)
    if missing:
        return None, jsonify({"ok": False, "error": "Debe completar: " + ", ".join(missing) + "."}), 400

    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        venc_expr = "CASE WHEN c.fecha_vencimiento IS NULL THEN NULL WHEN YEAR(c.fecha_vencimiento) < 1900 THEN DATE_ADD(c.fecha_vencimiento, INTERVAL 2000 YEAR) ELSE c.fecha_vencimiento END"

        query = f"""
            SELECT
                COALESCE(CAST(AES_DECRYPT(FROM_BASE64(p.asegurado), @SIS_KEY) AS CHAR), p.asegurado) AS asegurado,
                COALESCE(
                    CAST(AES_DECRYPT(FROM_BASE64(cl.direccion), @SIS_KEY) AS CHAR),
                    CAST(AES_DECRYPT(cl.direccion, @SIS_KEY) AS CHAR),
                    cl.direccion
                ) AS direccion,
                COALESCE(
                    CAST(AES_DECRYPT(FROM_BASE64(cl.telefono), @SIS_KEY) AS CHAR),
                    CAST(AES_DECRYPT(cl.telefono, @SIS_KEY) AS CHAR),
                    cl.telefono
                ) AS telefono,
                COALESCE(
                    CAST(AES_DECRYPT(FROM_BASE64(cl.razon_social), @SIS_KEY) AS CHAR),
                    CAST(AES_DECRYPT(cl.razon_social, @SIS_KEY) AS CHAR),
                    cl.razon_social
                ) AS contratante,
                COALESCE(
                    CAST(AES_DECRYPT(FROM_BASE64(p.poliza), @SIS_KEY) AS CHAR),
                    CAST(AES_DECRYPT(p.poliza, @SIS_KEY) AS CHAR),
                    p.poliza
                ) AS poliza,
                p.ejecutivo AS ejecutivo,
                p.cia,
                p.ramo AS ram,
                COALESCE(p.ramos_producto, p.ramo) AS prod,
                COALESCE(
                    CAST(AES_DECRYPT(FROM_BASE64(c.cupon), @SIS_KEY) AS CHAR),
                    CAST(AES_DECRYPT(c.cupon, @SIS_KEY) AS CHAR),
                    c.cupon
                ) AS cupon,
                ROW_NUMBER() OVER (
                    PARTITION BY p.idPoliza
                    ORDER BY ({venc_expr} IS NULL), {venc_expr} ASC, c.idCuota ASC
                ) AS num_cuota,
                {venc_expr} AS fec_vencimiento_cob,
                p.moneda AS mon,
                c.importe,
                c.fecha_pago AS fec_pago,
                c.factura,
                CASE
                    WHEN c.fecha_pago IS NULL AND {venc_expr} IS NOT NULL
                        THEN GREATEST(DATEDIFF(DATE(UTC_TIMESTAMP() - INTERVAL 12 HOUR), DATE({venc_expr})), 0)
                    ELSE 0
                END AS dias_vencidos,
                c.observacion AS ult_gestion,
                p.tipo_doc AS tp,
                p.vig_desde AS vig_del,
                p.vig_hasta AS vig_al,
                p.prima_comercial_igv AS prima_total,
                p.motivo,
                p.forma_pago AS tp_pago,
                NULL AS breve_descripcion
            FROM cuotas c
            LEFT JOIN (
                SELECT
                    MAX(idPoliza) AS idPoliza,
                    TRIM(
                        COALESCE(
                            CONVERT(AES_DECRYPT(FROM_BASE64(poliza), @SIS_KEY) USING utf8mb4),
                            poliza
                        ) COLLATE utf8mb4_0900_ai_ci
                    ) AS poliza_plain
                FROM polizas
                WHERE activo = 1
                  AND (anulado = 0 OR anulado IS NULL)
                  AND COALESCE(prima_anulada, 0) = 0
                GROUP BY poliza_plain
            ) p_lookup
              ON c.poliza_id IS NULL
             AND TRIM(
                    COALESCE(
                        CONVERT(AES_DECRYPT(FROM_BASE64(c.poliza), @SIS_KEY) USING utf8mb4),
                        c.poliza
                    ) COLLATE utf8mb4_0900_ai_ci
                 ) = p_lookup.poliza_plain
            INNER JOIN polizas p ON p.idPoliza = COALESCE(c.poliza_id, p_lookup.idPoliza)
            LEFT JOIN clientes cl ON p.cliente_id = cl.idCliente
            WHERE c.activo = 1
              AND p.activo = 1
              AND (p.anulado = 0 OR p.anulado IS NULL)
              AND COALESCE(p.prima_anulada, 0) = 0
        """

        params = []

        if fecha_desde:
            query += f" AND {venc_expr} >= %s"
            params.append(fecha_desde)
        if fecha_hasta:
            query += f" AND {venc_expr} < DATE_ADD(%s, INTERVAL 1 DAY)"
            params.append(fecha_hasta)
        if cliente_ids:
            query += " AND p.cliente_id IN (" + ",".join(["%s"] * len(cliente_ids)) + ")"
            params.extend(cliente_ids)
        if cias:
            query += " AND p.cia IN (" + ",".join(["%s"] * len(cias)) + ")"
            params.extend(cias)
        if ramos:
            query += " AND p.ramo IN (" + ",".join(["%s"] * len(ramos)) + ")"
            params.extend(ramos)
        if ejecutivos:
            query += " AND p.ejecutivo IN (" + ",".join(["%s"] * len(ejecutivos)) + ")"
            params.extend(ejecutivos)
        if sub_agentes:
            query += " AND p.sub_agente IN (" + ",".join(["%s"] * len(sub_agentes)) + ")"
            params.extend(sub_agentes)

        if poliza_cupon:
            like = f"%{poliza_cupon}%"
            poliza_expr = """
                TRIM(
                    COALESCE(
                        CONVERT(AES_DECRYPT(FROM_BASE64(p.poliza), @SIS_KEY) USING utf8mb4),
                        CONVERT(AES_DECRYPT(p.poliza, @SIS_KEY) USING utf8mb4),
                        p.poliza
                    ) COLLATE utf8mb4_0900_ai_ci
                )
            """
            cupon_expr = """
                TRIM(
                    COALESCE(
                        CONVERT(AES_DECRYPT(FROM_BASE64(c.cupon), @SIS_KEY) USING utf8mb4),
                        CONVERT(AES_DECRYPT(c.cupon, @SIS_KEY) USING utf8mb4),
                        c.cupon
                    ) COLLATE utf8mb4_0900_ai_ci
                )
            """
            query += f" AND ({poliza_expr} LIKE %s OR {cupon_expr} LIKE %s)"
            params.extend([like, like])

        estados_set = set([e for e in estados if e])
        if estados_set == {"PENDIENTE"}:
            query += " AND c.fecha_pago IS NULL"
        elif estados_set == {"PAGADO"}:
            query += " AND c.fecha_pago IS NOT NULL"

        role = session.get("role_name")
        user = session.get("user")
        if role == Roles.SUB_AGENTE and user:
            query += " AND (p.sub_agente = %s OR p.usuario_registro = %s)"
            params.extend([user, user])

        query += f" ORDER BY ({venc_expr} IS NULL), {venc_expr} ASC, c.idCuota DESC"

        cursor.execute(query, tuple(params))
        rows = cursor.fetchall() or []

        for r in rows:
            if r.get("fec_vencimiento_cob"):
                try:
                    r["fec_vencimiento_cob"] = r["fec_vencimiento_cob"].strftime("%Y-%m-%d")
                except Exception:
                    r["fec_vencimiento_cob"] = str(r["fec_vencimiento_cob"])[:10]
            if r.get("fec_pago"):
                try:
                    r["fec_pago"] = r["fec_pago"].strftime("%Y-%m-%d")
                except Exception:
                    r["fec_pago"] = str(r["fec_pago"])[:10]
            if r.get("vig_del"):
                try:
                    r["vig_del"] = r["vig_del"].strftime("%Y-%m-%d")
                except Exception:
                    r["vig_del"] = str(r["vig_del"])[:10]
            if r.get("vig_al"):
                try:
                    r["vig_al"] = r["vig_al"].strftime("%Y-%m-%d")
                except Exception:
                    r["vig_al"] = str(r["vig_al"])[:10]

        return rows, None, None
    except Exception as e:
        return None, jsonify({"ok": False, "error": str(e)}), 500
    finally:
        try:
            if cursor:
                cursor.close()
        except Exception:
            pass
        try:
            if conn:
                conn.close()
        except Exception:
            pass


@bp.route("/api/cobranzas/estado-cuenta-cupones", methods=["GET"])
def api_cobranzas_estado_cuenta_cupones():
    if "user" not in session:
        return {"ok": False, "error": "No autenticado"}, 401

    rows, err_resp, err_status = _fetch_rows()
    if err_resp is not None:
        return err_resp, err_status
    return jsonify({"ok": True, "rows": rows})


@bp.route("/api/cobranzas/estado-cuenta-cupones/export/xlsx", methods=["GET"])
def api_cobranzas_estado_cuenta_cupones_export_xlsx():
    if "user" not in session:
        return {"ok": False, "error": "No autenticado"}, 401

    rows, err_resp, err_status = _fetch_rows()
    if err_resp is not None:
        return err_resp, err_status

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Cupones"

    headers = [
        "ASEGURADO",
        "DIRECCION",
        "TELEFONO",
        "CONTRATANTE",
        "POLIZA",
        "EJECUTIVO",
        "CIA",
        "RAM",
        "PROD",
        "CUPON",
        "NUM_CUOTA",
        "FEC_VENCIMIENTO COB",
        "MON",
        "IMPORTE",
        "FEC_PAGO",
        "FACTURA",
        "DIAS_VENCIDOS",
        "ULT_GESTION",
        "TP",
        "VIG_DEL",
        "VIG_AL",
        "PRIMA_TOTAL",
        "MOTIVO",
        "TP_PAGO",
        "BREVE_DESCRIPCION",
    ]
    fecha_desde = (request.args.get("fecha_desde") or "").strip()
    fecha_hasta = (request.args.get("fecha_hasta") or "").strip()
    title = "ESTADO DE CUENTA DE CUPONES"
    if fecha_desde and fecha_hasta:
        title = f"{title} — {fecha_desde} a {fecha_hasta}"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F59A3")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    header_fill = PatternFill("solid", fgColor="399AD6")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 22

    money_cols = {14, 22}
    int_cols = {11, 17}

    for i, r in enumerate(rows, start=3):
        row_vals = [
            r.get("asegurado") or "",
            r.get("direccion") or "",
            r.get("telefono") or "",
            r.get("contratante") or "",
            r.get("poliza") or "",
            r.get("ejecutivo") or "",
            r.get("cia") or "",
            r.get("ram") or "",
            r.get("prod") or "",
            r.get("cupon") or "",
            int(r.get("num_cuota") or 0),
            r.get("fec_vencimiento_cob") or "",
            r.get("mon") or "",
            float(r.get("importe") or 0),
            r.get("fec_pago") or "",
            r.get("factura") or "",
            int(r.get("dias_vencidos") or 0),
            r.get("ult_gestion") or "",
            r.get("tp") or "",
            r.get("vig_del") or "",
            r.get("vig_al") or "",
            float(r.get("prima_total") or 0),
            r.get("motivo") or "",
            r.get("tp_pago") or "",
            r.get("breve_descripcion") or "",
        ]

        fill = PatternFill("solid", fgColor="F4F8FF") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.border = border
            cell.fill = fill
            cell.font = Font(size=9)
            if col_idx in money_cols:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in int_cols:
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    total_row = len(rows) + 3
    ws.cell(row=total_row, column=13, value="TOTAL").font = Font(bold=True, size=9)
    tot_importe_cell = ws.cell(row=total_row, column=14)
    tot_prima_cell = ws.cell(row=total_row, column=22)
    for c in (tot_importe_cell, tot_prima_cell):
        c.font = Font(bold=True, size=9)
        c.number_format = "#,##0.00"
        c.alignment = Alignment(horizontal="right", vertical="center")
    if rows:
        tot_importe_cell.value = f"=SUBTOTAL(109,N3:N{total_row-1})"
        tot_prima_cell.value = f"=SUBTOTAL(109,V3:V{total_row-1})"
    else:
        tot_importe_cell.value = 0
        tot_prima_cell.value = 0

    col_widths = [
        22,
        22,
        14,
        28,
        16,
        18,
        14,
        14,
        18,
        18,
        10,
        16,
        10,
        14,
        14,
        16,
        12,
        18,
        8,
        12,
        12,
        14,
        20,
        14,
        24,
    ]
    for i, w in enumerate(col_widths[: len(headers)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    upload_folder = current_app.config.get("UPLOAD_FOLDER", os.path.join(current_app.root_path, "uploads"))
    exports_dir = os.path.join(upload_folder, "exports")
    os.makedirs(exports_dir, exist_ok=True)

    from datetime import datetime

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"estado_cuenta_cupones_{ts}.xlsx"
    filepath = os.path.join(exports_dir, filename)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True, download_name=filename)

